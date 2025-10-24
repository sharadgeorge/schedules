import streamlit as st
import sys
import os
from pathlib import Path
import tempfile

# Configure page
st.set_page_config(
    page_title="Cardiology - Schedule Management",
    page_icon="❤️",
    layout="centered",
    initial_sidebar_state="expanded"
)

# Hide Streamlit's automatic page navigation
st.markdown("""
    <style>
        [data-testid="stSidebarNav"] {
            display: none;
        }
    </style>
""", unsafe_allow_html=True)

# Sidebar with About section and explicit navigation
with st.sidebar:
    st.markdown("### About")
    st.info("**Project by JA RAD**")
    st.markdown("---")
    
    # Explicit navigation buttons
    st.markdown("### Navigation")
    
    # Button to navigate to Radiology
    if st.button("🔬 Go to Radiology", use_container_width=True):
        st.switch_page("app.py")
    
    # Current page indicator
    st.markdown("**❤️ Cardiology** (Current)")

# Main page title
st.title("❤️ Cardiology Schedule Management")
st.markdown("---")

# Section: Convert Cardiology Schedules for Import
st.header("🔄 Convert Cardiology Schedules for Import")
st.markdown("Upload the required Cardiology Schedule Excel files to generate the import CSV file.")
st.info("ℹ️ Currently supports 2 input files. This will be expandable in the future for additional teams.")

# File uploaders
st.subheader("Upload Schedule Files (in order):")

col1, col2 = st.columns(2)

with col1:
    cardio_file1 = st.file_uploader(
        "1. Cardiovascular Schedule (Excel)", 
        type=['xlsx'],
        key='cardio_file1',
        help="Upload the Cardiovascular schedule Excel file"
    )

with col2:
    cardio_file2 = st.file_uploader(
        "2. Interventional Cardiologist Schedule (Excel)", 
        type=['xlsx'],
        key='cardio_file2',
        help="Upload the Interventional Cardiologist schedule Excel file"
    )

# Future expandability note
with st.expander("ℹ️ About Future Expansion"):
    st.markdown("""
    This converter currently processes:
    - **Team 8**: Cardiovascular (Echo Tech Adult & Pediatric)
    - **Team 94**: Interventional Cardiologist
    
    In the future, additional file upload slots can be added here for new teams.
    The converter script will need to be updated accordingly to handle the new team configurations.
    """)

# Process files when both are uploaded
if cardio_file1 and cardio_file2:
    if st.button("Convert to Import Format", key='convert_cardio'):
        with st.spinner("Converting Cardiology schedules to import format..."):
            try:
                # Create temporary directory for processing
                with tempfile.TemporaryDirectory() as temp_dir:
                    # Save uploaded files
                    file1_path = Path(temp_dir) / cardio_file1.name
                    file2_path = Path(temp_dir) / cardio_file2.name
                    
                    with open(file1_path, 'wb') as f:
                        f.write(cardio_file1.getbuffer())
                    
                    with open(file2_path, 'wb') as f:
                        f.write(cardio_file2.getbuffer())
                    
                    # Import necessary modules
                    sys.path.insert(0, str(Path(__file__).parent.parent))
                    import oncall_converter_Cardiology_demo_v3 as cardio_converter
                    
                    # Load workbooks
                    import openpyxl
                    wb_cardio = openpyxl.load_workbook(file1_path, data_only=True)
                    wb_intv = openpyxl.load_workbook(file2_path, data_only=True)
                    
                    # Extract month and year from Cardiovascular file
                    import calendar
                    from datetime import datetime
                    
                    # Try to find the sheet with "On Call" or "On-Call" in the name
                    cardio_sheet = None
                    for sheet_name in wb_cardio.sheetnames:
                        if 'on' in sheet_name.lower() and 'call' in sheet_name.lower():
                            cardio_sheet = wb_cardio[sheet_name]
                            break
                    
                    if not cardio_sheet:
                        cardio_sheet = wb_cardio.active
                    
                    # Extract month from cell B4
                    month_num, year = cardio_converter.extract_month_year_from_file(wb_cardio, cardio_sheet.title, 'B4')
                    
                    if month_num is None:
                        # Try to extract from filename
                        filename_lower = file1_path.stem.lower()
                        for m in range(1, 13):
                            if calendar.month_name[m].lower() in filename_lower:
                                month_num = m
                                break
                        
                        if month_num is None:
                            month_num = datetime.now().month
                        
                        if year is None:
                            year = datetime.now().year
                    
                    month_name = calendar.month_name[month_num]
                    
                    # Read data from both files
                    cardiovascular_data = cardio_converter.read_cardiovascular_data(wb_cardio, month_num, year)
                    interventional_data = cardio_converter.read_interventional_data(wb_intv, month_num, year)
                    
                    # Create output data
                    output_data = cardio_converter.create_output_data(cardiovascular_data, interventional_data, year, month_num)
                    
                    # Write CSV output
                    import csv
                    csv_output = Path(temp_dir) / "Import_OnCall_Cardiology.csv"
                    with open(csv_output, 'w', newline='') as csvfile:
                        fieldnames = ['EMPLOYEE', 'TEAM', 'STARTDATE', 'STARTTIME', 
                                     'ENDDATE', 'ENDTIME', 'ROLE', 'NOTES', 'ORDER', 'TEAMCOMMENT']
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter='^')
                        writer.writeheader()
                        writer.writerows(output_data)
                    
                    # Read the CSV file for download
                    with open(csv_output, 'rb') as f:
                        csv_data = f.read()
                    
                    # Success message with statistics
                    cardio_days_with_assignments = len([d for d in cardiovascular_data if cardiovascular_data[d]])
                    intv_days_with_assignments = len(interventional_data)
                    
                    st.success(f"✅ Successfully generated {len(output_data)} schedule entries for {month_name} {year}")
                    
                    # Show some statistics
                    with st.expander("📊 Processing Statistics"):
                        st.markdown(f"""
                        - **Month/Year**: {month_name} {year}
                        - **Total Entries**: {len(output_data)}
                        - **Cardiovascular Assignments**: {cardio_days_with_assignments} days
                        - **Interventional Assignments**: {intv_days_with_assignments} days
                        - **Expected Entries**: {calendar.monthrange(year, month_num)[1] * 3} (3 per day)
                        """)
                    
                    # Download button
                    st.download_button(
                        label="📥 Download Import_OnCall_Cardiology.csv",
                        data=csv_data,
                        file_name="Import_OnCall_Cardiology.csv",
                        mime="text/csv",
                        key='download_cardio_csv'
                    )
                    
            except Exception as e:
                st.error(f"❌ An error occurred: {str(e)}")
                st.error("Please ensure you've uploaded valid Excel files with the correct format.")
                
                # Show error details in expander
                with st.expander("🔍 Error Details"):
                    import traceback
                    st.code(traceback.format_exc())

else:
    st.info("👆 Please upload both required Excel files to begin conversion.")

# Footer
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: gray; font-size: 0.8em;'>"
    "Schedule Management System | Powered by Streamlit"
    "</div>",
    unsafe_allow_html=True
)
