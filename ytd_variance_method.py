# NEW METHOD: Calculate YTD Variance
# Add this method to your OnCallScheduler class

def calculate_ytd_variance(self, ws_read=None):
    """
    Calculate variance between actual YTD totals and target YTD totals.
    Reads from columns AM/AN/AO (actual) and AQ/AR/AS (target).
    
    Returns dict with variance metrics for display
    """
    import openpyxl
    
    if ws_read is None:
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws_read = wb['Sheet1']
    
    print("\n=== Calculating YTD Variance ===")
    
    # Column mappings (1-indexed)
    actual_cols = {
        'weekday': 39,  # Column AM (weekdays/wd)
        'thu': 40,      # Column AN (thursdays)
        'weekend': 41   # Column AO (weekends)
    }
    
    target_cols = {
        'weekday': 43,  # Column AQ (weekdays/wd)
        'thu': 44,      # Column AR (thursdays)
        'weekend': 45   # Column AS (weekends)
    }
    
    # Initialize variance tracking
    variance_data = {}
    
    # GEN section rows
    gen_rads = ['NN', 'MB', 'LK', 'PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'IG', 'MF', 'AS']
    gen_start_row = 5  # Assuming GEN section starts at row 5
    
    print("\n--- GEN Section Variance ---")
    for idx, rad in enumerate(gen_rads):
        row = gen_start_row + idx
        variance_data[f'GEN_{rad}'] = {}
        
        for day_type in ['weekday', 'thu', 'weekend']:
            actual_cell = ws_read.cell(row, actual_cols[day_type])
            target_cell = ws_read.cell(row, target_cols[day_type])
            
            actual_val = actual_cell.value if actual_cell.value is not None else 0
            target_val = target_cell.value if target_cell.value is not None else 0
            
            # Handle potential string/formula values
            try:
                actual_val = float(actual_val) if actual_val != '' else 0
                target_val = float(target_val) if target_val != '' else 0
            except:
                actual_val = 0
                target_val = 0
            
            variance = actual_val - target_val
            variance_data[f'GEN_{rad}'][day_type] = {
                'actual': actual_val,
                'target': target_val,
                'variance': variance
            }
            
            if abs(variance) > 0.5:  # Only show significant variances
                print(f"  {rad} {day_type}: actual={actual_val:.1f}, target={target_val:.1f}, variance={variance:+.1f}")
    
    # IRA section rows
    ira_rads = ['IG', 'MF', 'AS']
    ira_start_row = 20  # Assuming IRA section starts at row 20
    
    print("\n--- IRA Section Variance ---")
    for idx, rad in enumerate(ira_rads):
        row = ira_start_row + idx
        variance_data[f'IRA_{rad}'] = {}
        
        for day_type in ['weekday', 'thu', 'weekend']:
            actual_cell = ws_read.cell(row, actual_cols[day_type])
            target_cell = ws_read.cell(row, target_cols[day_type])
            
            actual_val = actual_cell.value if actual_cell.value is not None else 0
            target_val = target_cell.value if target_cell.value is not None else 0
            
            try:
                actual_val = float(actual_val) if actual_val != '' else 0
                target_val = float(target_val) if target_val != '' else 0
            except:
                actual_val = 0
                target_val = 0
            
            variance = actual_val - target_val
            variance_data[f'IRA_{rad}'][day_type] = {
                'actual': actual_val,
                'target': target_val,
                'variance': variance
            }
            
            if abs(variance) > 0.5:
                print(f"  {rad} {day_type}: actual={actual_val:.1f}, target={target_val:.1f}, variance={variance:+.1f}")
    
    # MRI section rows
    mri_rads = ['PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'MF', 'AS']
    mri_start_row = 26  # Assuming MRI section starts at row 26
    
    print("\n--- MRI Section Variance ---")
    for idx, rad in enumerate(mri_rads):
        row = mri_start_row + idx
        variance_data[f'MRI_{rad}'] = {}
        
        for day_type in ['weekday', 'thu', 'weekend']:
            actual_cell = ws_read.cell(row, actual_cols[day_type])
            target_cell = ws_read.cell(row, target_cols[day_type])
            
            actual_val = actual_cell.value if actual_cell.value is not None else 0
            target_val = target_cell.value if target_cell.value is not None else 0
            
            try:
                actual_val = float(actual_val) if actual_val != '' else 0
                target_val = float(target_val) if target_val != '' else 0
            except:
                actual_val = 0
                target_val = 0
            
            variance = actual_val - target_val
            variance_data[f'MRI_{rad}'][day_type] = {
                'actual': actual_val,
                'target': target_val,
                'variance': variance
            }
            
            if abs(variance) > 0.5:
                print(f"  {rad} {day_type}: actual={actual_val:.1f}, target={target_val:.1f}, variance={variance:+.1f}")
    
    # Calculate aggregate variance metrics
    aggregate_variance = {
        'weekday': {'sum_abs': 0, 'sum_sq': 0, 'count': 0},
        'thu': {'sum_abs': 0, 'sum_sq': 0, 'count': 0},
        'weekend': {'sum_abs': 0, 'sum_sq': 0, 'count': 0}
    }
    
    for rad_key, day_types in variance_data.items():
        for day_type in ['weekday', 'thu', 'weekend']:
            if day_type in day_types:
                variance = day_types[day_type]['variance']
                aggregate_variance[day_type]['sum_abs'] += abs(variance)
                aggregate_variance[day_type]['sum_sq'] += variance ** 2
                aggregate_variance[day_type]['count'] += 1
    
    # Calculate summary statistics
    print("\n=== Aggregate Variance Summary ===")
    print(f"{'Type':<10} {'Total Abs Var':<15} {'Avg Abs Var':<15} {'RMSE':<15}")
    print("-" * 60)
    
    summary = {}
    for day_type in ['weekday', 'thu', 'weekend']:
        total_abs = aggregate_variance[day_type]['sum_abs']
        count = aggregate_variance[day_type]['count']
        avg_abs = total_abs / count if count > 0 else 0
        rmse = (aggregate_variance[day_type]['sum_sq'] / count) ** 0.5 if count > 0 else 0
        
        summary[day_type] = {
            'total_abs_variance': total_abs,
            'avg_abs_variance': avg_abs,
            'rmse': rmse
        }
        
        print(f"{day_type:<10} {total_abs:<15.2f} {avg_abs:<15.3f} {rmse:<15.3f}")
    
    # Overall quality score (lower is better)
    # Weighted: weekends = 3x, thursdays = 2x, weekdays = 1x
    overall_score = (
        summary['weekend']['rmse'] * 3 +
        summary['thu']['rmse'] * 2 +
        summary['weekday']['rmse'] * 1
    ) / 6
    
    print(f"\nOverall Quality Score (RMSE weighted): {overall_score:.3f}")
    print("  (Lower is better - weights: weekend=3x, thu=2x, weekday=1x)")
    
    return {
        'variance_data': variance_data,
        'aggregate': aggregate_variance,
        'summary': summary,
        'overall_score': overall_score
    }


def format_variance_for_display(variance_results):
    """
    Format variance results for nice display in Streamlit or console.
    Returns formatted string.
    """
    output = []
    
    output.append("\n" + "="*70)
    output.append("YTD VARIANCE ANALYSIS (Actual vs Target)")
    output.append("="*70)
    
    summary = variance_results['summary']
    
    output.append("\nAggregate Variance by Day Type:")
    output.append(f"{'Type':<12} {'Total Abs Var':<16} {'Avg Abs Var':<16} {'RMSE':<12}")
    output.append("-" * 70)
    
    for day_type in ['weekend', 'thu', 'weekday']:
        total_abs = summary[day_type]['total_abs_variance']
        avg_abs = summary[day_type]['avg_abs_variance']
        rmse = summary[day_type]['rmse']
        
        # Add emoji indicators
        if rmse < 1.0:
            indicator = "✓"
        elif rmse < 2.0:
            indicator = "⚠"
        else:
            indicator = "⚠⚠"
        
        output.append(
            f"{day_type:<12} {total_abs:<16.2f} {avg_abs:<16.3f} {rmse:<12.3f} {indicator}"
        )
    
    overall_score = variance_results['overall_score']
    output.append("\n" + "-" * 70)
    output.append(f"Overall Quality Score: {overall_score:.3f}")
    output.append("  (Weighted RMSE: weekend=3x, thu=2x, weekday=1x)")
    
    if overall_score < 1.5:
        output.append("  ✓ EXCELLENT balance!")
    elif overall_score < 2.5:
        output.append("  ✓ GOOD balance")
    elif overall_score < 3.5:
        output.append("  ⚠ FAIR balance (could improve)")
    else:
        output.append("  ⚠ POOR balance (needs improvement)")
    
    output.append("="*70)
    
    return "\n".join(output)
