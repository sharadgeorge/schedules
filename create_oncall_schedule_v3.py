"""
OnCall Schedule Generator for Radiology Department
Handles GEN, IRA, and MRI call assignments with intelligent load balancing

CRITICAL NOTES:
- Day 1 always starts at Column D in Excel (formula: openpyxl_column = day + 3)
- GEN pairs: Thursday + Saturday (same rad)
- IRA triplets: Thursday + Friday + Saturday (same rad)
- Max 3 weekend days per rad per month (GEN only)
- Max 7 total GEN days per month
- Max 12 total IRA days per month
- Balances toward YTD targets while respecting monthly limits
"""

import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import calendar

# Configuration
MAX_MONTHLY_GEN = 7
MAX_MONTHLY_IRA = 12
MAX_MONTHLY_MRI = 8
MAX_MONTHLY_WEEKENDS_GEN = 3  # Hard limit for GEN only
CONSECUTIVE_WEEKEND_PENALTY = 100
SOFT_CONSTRAINT_PENALTY = 300  # Penalty for soft constraints

# Employee start dates
START_DATES = {
    'NN': (7, 15), 'MB': (1, 1), 'LK': (1, 1), 'PR': (7, 15),
    'AT': (1, 1), 'AK': (1, 1), 'MC': (1, 1), 'AO': (1, 1),
    'MM': (1, 1), 'IG': (5, 1), 'MF': (1, 1), 'AS': (1, 1)
}

# Section definitions
GEN_RADS = ['NN', 'MB', 'LK', 'PR', 'AT', 'AK', 'MC', 'AO', 'MM']
GEN_RADS_WITH_IRA = ['NN', 'MB', 'LK', 'PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'IG', 'MF', 'AS']
IRA_RADS = ['IG', 'MF', 'AS']
MRI_RADS = ['PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'MF', 'AS']

# Excel row mappings
GEN_ROWS = {
    'TA': 5, 'NN': 7, 'MB': 8, 'LK': 9, 'PR': 10, 'AT': 11,
    'AK': 12, 'MC': 13, 'AO': 14, 'MM': 15, 'IG': 17, 'MF': 18, 'AS': 19
}
IRA_ROWS = {'IG': 24, 'MF': 25, 'AS': 26}
MRI_ROWS = {'PR': 30, 'AT': 31, 'AK': 32, 'MC': 33, 'AO': 34, 'MM': 35, 'MF': 36, 'AS': 37}

# Column indices - CORRECTED (openpyxl uses 1-based indexing)
COL_AM = 39  # YTD weekday (Column AM)
COL_AN = 40  # YTD Thursday (Column AN)
COL_AO = 41  # YTD weekend (Column AO)
# COL_AP = 42 is BLANK - never write to this column
COL_AQ = 43  # YTD Target weekday (Column AQ)
COL_AR = 44  # YTD Target Thursday (Column AR)
COL_AS = 45  # YTD Target weekend (Column AS)


class OnCallScheduler:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.ws = self.wb['Sheet1']
        
        self.month, self.year = self.extract_month_year()
        self.days_in_month = calendar.monthrange(self.year, self.month)[1]
        
        self.ytd_cache = self.cache_ytd_totals()
        self.ytd_targets = self.calculate_ytd_targets()
        
        self.monthly_counts = defaultdict(lambda: {'weekday': 0, 'thu': 0, 'weekend': 0})
        self.gen_monthly_total = defaultdict(int)
        self.ira_monthly_total = defaultdict(int)
        self.mri_monthly_total = defaultdict(int)
        
        self.assignments = {'GEN': {}, 'IRA': {}, 'MRI': {}}
        
        self.locked_assignments = self.identify_locked_assignments()
        self.vacation_days = self.identify_vacation_days()
        self.special_requests_off = {}  # Hard constraints (cannot assign)
        self.soft_constraints_off = {}  # Soft constraints (penalize but can assign)


    def get_user_preferences(self):
        """Prompt user for scheduling preferences and constraints"""
        print("\n" + "="*80)
        print("SCHEDULING PREFERENCES AND CONSTRAINTS")
        print("="*80)
        print(f"Month: {calendar.month_name[self.month]} {self.year}")
        print("\nPlease provide scheduling preferences to optimize the schedule.")
        print("Press Enter to skip any question.\n")
        
        # Initialize dictionaries
        self.special_requests_off = defaultdict(set)  # Hard constraints
        self.soft_constraints_off = defaultdict(set)   # Soft constraints
        
        # 1. GEN rads on call last weekends of prior month (HARD constraint for first weekend only)
        print("-" * 80)
        print("1. GEN RADS ON CALL LAST WEEKENDS OF PRIOR MONTH")
        print("   (Hard constraint: These rads will NOT be assigned first WEEKEND of current month)")
        print("   (Can assign weekdays if at least 1 day gap from last weekend)")
        print("-" * 80)
        
        last_weekend_gen = []
        for i in range(2):
            while True:
                rad = input(f"   GEN rad #{i+1} on last weekend of prior month (or Enter to skip): ").strip().upper()
                if not rad:
                    break
                if rad in GEN_RADS_WITH_IRA:
                    last_weekend_gen.append(rad)
                    break
                else:
                    print(f"   ERROR: '{rad}' is not a valid GEN rad. Valid: {', '.join(GEN_RADS_WITH_IRA)}")
        
        # Find first weekend (Friday + Saturday) of current month
        first_weekend_days = set()
        for day in range(1, min(15, self.days_in_month + 1)):  # Check first 2 weeks
            date = datetime(self.year, self.month, day)
            if date.weekday() == 4:  # Friday
                first_weekend_days.add(day)
                if day + 1 <= self.days_in_month:
                    first_weekend_days.add(day + 1)  # Saturday
                break
        
        # Calculate when prior month ended
        first_day_of_month = datetime(self.year, self.month, 1)
        last_day_of_prior_month = first_day_of_month - timedelta(days=1)
        days_since_prior_month_end = {day: day for day in range(1, self.days_in_month + 1)}
        
        for rad in last_weekend_gen:
            # Hard constraint: Block first weekend
            self.special_requests_off[rad].update(first_weekend_days)
            
            # Hard constraint: Block day 1 if prior month ended on weekend (no gap)
            if last_day_of_prior_month.weekday() in [4, 5]:  # Fri or Sat
                # If prior month ended on Fri/Sat, they likely worked until then
                # So day 1 of current month has 0 days gap - block it
                if 1 <= self.days_in_month:
                    self.special_requests_off[rad].add(1)
            
            weekend_str = ', '.join(map(str, sorted(first_weekend_days))) if first_weekend_days else 'none'
            gap_note = " and day 1 (no gap)" if last_day_of_prior_month.weekday() in [4, 5] else ""
            print(f"   → {rad} will NOT be assigned GEN on first weekend (days {weekend_str}){gap_note}")
            print(f"      (Can be assigned weekdays with at least 1 day gap)")
        
        # 2. GEN rad on call last day of prior month (HARD constraint for day 1-2)
        print("\n" + "-" * 80)
        print("2. GEN RAD ON CALL LAST DAY OF PRIOR MONTH")
        print("   (Hard constraint: This rad will NOT be assigned GEN on days 1-2)")
        print("-" * 80)
        
        while True:
            last_day_gen = input("   GEN rad on last day of prior month (or Enter to skip): ").strip().upper()
            if not last_day_gen:
                break
            if last_day_gen in GEN_RADS_WITH_IRA:
                # Only add if this is after the last weekend rads (to avoid duplication)
                if last_day_gen not in last_weekend_gen:
                    self.special_requests_off[last_day_gen].update([1, 2])
                    print(f"   → {last_day_gen} will NOT be assigned GEN on days 1-2")
                else:
                    print(f"   → {last_day_gen} already constrained from last weekend")
                break
            else:
                print(f"   ERROR: '{last_day_gen}' is not a valid GEN rad. Valid: {', '.join(GEN_RADS_WITH_IRA)}")
        
        # 3. IRA rad on call last weekend of prior month (SOFT constraint for first week)
        print("\n" + "-" * 80)
        print("3. IRA RAD ON CALL LAST WEEKEND OF PRIOR MONTH")
        print("   (Soft constraint: This rad is DISCOURAGED but can be assigned during first week)")
        print("-" * 80)
        
        while True:
            last_weekend_ira = input("   IRA rad on last weekend of prior month (or Enter to skip): ").strip().upper()
            if not last_weekend_ira:
                break
            if last_weekend_ira in IRA_RADS:
                # Soft constraint for first week
                first_week_days = set(range(1, min(8, self.days_in_month + 1)))
                self.soft_constraints_off[last_weekend_ira] = first_week_days.copy()
                print(f"   → {last_weekend_ira} is DISCOURAGED (but can be assigned) for IRA during days 1-7")
                break
            else:
                print(f"   ERROR: '{last_weekend_ira}' is not a valid IRA rad. Valid: {', '.join(IRA_RADS)}")
        
        # 4. Additional preference requests
        print("\n" + "-" * 80)
        print("4. ADDITIONAL PREFERENCE REQUESTS (OFF)")
        print("   Enter additional days when specific rads prefer NOT to be on call")
        print("   Format: RAD SECTION DAY (e.g., 'MB GEN 15' or 'IG IRA 20')")
        print("   Type 'done' when finished")
        print("-" * 80)
        
        while True:
            request = input("   Enter request (or 'done'): ").strip()
            if request.lower() == 'done' or not request:
                break
            
            parts = request.upper().split()
            if len(parts) != 3:
                print("   ERROR: Format should be 'RAD SECTION DAY' (e.g., 'MB GEN 15')")
                continue
            
            rad, section, day_str = parts
            
            # Validate section
            if section not in ['GEN', 'IRA']:
                print(f"   ERROR: Section must be 'GEN' or 'IRA', not '{section}'")
                continue
            
            # Validate rad for section
            if section == 'GEN' and rad not in GEN_RADS_WITH_IRA:
                print(f"   ERROR: '{rad}' is not a valid GEN rad. Valid: {', '.join(GEN_RADS_WITH_IRA)}")
                continue
            if section == 'IRA' and rad not in IRA_RADS:
                print(f"   ERROR: '{rad}' is not a valid IRA rad. Valid: {', '.join(IRA_RADS)}")
                continue
            
            # Validate day
            try:
                day = int(day_str)
                if day < 1 or day > self.days_in_month:
                    print(f"   ERROR: Day must be between 1 and {self.days_in_month}")
                    continue
            except ValueError:
                print(f"   ERROR: '{day_str}' is not a valid day number")
                continue
            
            # Ask if hard or soft constraint
            while True:
                constraint_type = input(f"   Is this a HARD constraint (cannot assign) or SOFT (discouraged)? [H/S]: ").strip().upper()
                if constraint_type in ['H', 'HARD']:
                    self.special_requests_off[rad].add(day)
                    print(f"   → {rad} will NOT be assigned {section} on day {day} (HARD)")
                    break
                elif constraint_type in ['S', 'SOFT']:
                    if rad not in self.soft_constraints_off:
                        self.soft_constraints_off[rad] = set()
                    self.soft_constraints_off[rad].add(day)
                    print(f"   → {rad} is DISCOURAGED for {section} on day {day} (SOFT)")
                    break
                else:
                    print("   Please enter 'H' for hard or 'S' for soft")
        
        # Add post-vacation soft constraints
        print("\n" + "-" * 80)
        print("5. POST-VACATION SOFT CONSTRAINTS")
        print("   Adding day AFTER vacation as soft constraint (discouraged but can assign)...")
        print("-" * 80)
        
        for rad, vacation_days in self.vacation_days.items():
            for vac_day in vacation_days:
                day_after = vac_day + 1
                if day_after <= self.days_in_month:
                    if rad not in self.soft_constraints_off:
                        self.soft_constraints_off[rad] = set()
                    self.soft_constraints_off[rad].add(day_after)
                    print(f"   → {rad} is DISCOURAGED on day {day_after} (day after vacation day {vac_day})")
        
        # Summary
        print("\n" + "="*80)
        print("CONSTRAINTS SUMMARY")
        print("="*80)
        
        print("\nHARD CONSTRAINTS (Cannot assign):")
        print("  - Vacation days (blue cells)")
        print("  - Day before vacation")
        if self.special_requests_off:
            for rad, days in sorted(self.special_requests_off.items()):
                if days:
                    days_str = ', '.join(map(str, sorted(days)))
                    print(f"  - {rad}: days {days_str}")
        else:
            print("  - No additional hard constraints")
        
        print("\nSOFT CONSTRAINTS (Discouraged, can assign if necessary):")
        print("  - Day after vacation")
        if self.soft_constraints_off:
            for rad, days in sorted(self.soft_constraints_off.items()):
                if days:
                    days_str = ', '.join(map(str, sorted(days)))
                    print(f"  - {rad}: days {days_str}")
        else:
            print("  - No additional soft constraints")
        
        print("\nON REQUESTS (Locked - X marks in Excel):")
        for section in ['GEN', 'IRA', 'MRI']:
            if self.locked_assignments[section]:
                print(f"  {section}:")
                for day, rad in sorted(self.locked_assignments[section].items()):
                    print(f"    Day {day}: {rad}")
        
        print("\n" + "="*80)
        input("Press Enter to continue with schedule generation...")
        print()
    
    def calculate_ytd_targets(self):
        """Calculate YTD targets based on total working days through end of current month"""
        targets = {}
        
        holiday_ranges = [
            (datetime(self.year, 2, 23), datetime(self.year, 2, 23)),  # Feb 23 (Sunday)
            (datetime(self.year, 3, 27), datetime(self.year, 4, 5)),   # Spring holidays
            (datetime(self.year, 6, 5), datetime(self.year, 6, 14)),   # Summer holidays
            (datetime(self.year, 9, 23), datetime(self.year, 9, 23))   # Sep 23 (Tuesday)
        ]
        
        def is_holiday(date):
            for start, end in holiday_ranges:
                if start <= date <= end:
                    return True
            return False
        
        def get_day_type_from_date(date):
            weekday = date.weekday()
            if weekday == 3:
                return 'thu'
            elif weekday in [4, 5]:
                return 'weekend'
            else:
                return 'weekday'
        
        def count_rad_available_days_ytd(start_month, start_day):
            counts = {'weekday': 0, 'thu': 0, 'weekend': 0}
            
            for month in range(start_month, self.month + 1):
                days_in_month = calendar.monthrange(self.year, month)[1]
                start_of_month = start_day if month == start_month else 1
                
                for day in range(start_of_month, days_in_month + 1):
                    date = datetime(self.year, month, day)
                    if not is_holiday(date):
                        day_type = get_day_type_from_date(date)
                        counts[day_type] += 1
            
            return counts
        
        # Count total working days from Jan 1 through end of current month
        total_working_days_ytd = count_rad_available_days_ytd(1, 1)
        
        # GEN section (only the 9 primary GEN rads: NN, MB, LK, PR, AT, AK, MC, AO, MM)
        # Adjust for AS covering 1 weekend day in February (reduce total GEN workload by 1 weekend)
        gen_total_working_days = total_working_days_ytd.copy()
        gen_total_working_days['weekend'] -= 1  # AS covered 1 GEN weekend in Feb
        
        gen_available_days = {}
        gen_total_availability_shares = {'weekday': 0, 'thu': 0, 'weekend': 0}
        
        for rad in GEN_RADS:
            start_month, start_day = START_DATES[rad]
            available = count_rad_available_days_ytd(start_month, start_day)
            gen_available_days[rad] = available
            
            # Calculate share (fraction) for each rad based on adjusted GEN total
            for day_type in ['weekday', 'thu', 'weekend']:
                share = available[day_type] / gen_total_working_days[day_type]
                gen_total_availability_shares[day_type] += share
        
        # Calculate targets: Total_days / Total_shares * Rad_share
        for rad in GEN_RADS:
            targets[rad] = {}
            for day_type in ['weekday', 'thu', 'weekend']:
                if gen_total_availability_shares[day_type] > 0:
                    rad_share = gen_available_days[rad][day_type] / gen_total_working_days[day_type]
                    targets[rad][day_type] = (gen_total_working_days[day_type] / gen_total_availability_shares[day_type]) * rad_share
                else:
                    targets[rad][day_type] = 0.0
        
        # IRA section (IG, MF, AS) - use full total working days (no adjustment)
        ira_available_days = {}
        ira_total_availability_shares = {'weekday': 0, 'thu': 0, 'weekend': 0}
        
        for rad in IRA_RADS:
            start_month, start_day = START_DATES[rad]
            available = count_rad_available_days_ytd(start_month, start_day)
            ira_available_days[rad] = available
            
            # Calculate share (fraction) for each rad
            for day_type in ['weekday', 'thu', 'weekend']:
                share = available[day_type] / total_working_days_ytd[day_type]
                ira_total_availability_shares[day_type] += share
        
        # Calculate targets: Total_days / Total_shares * Rad_share
        for rad in IRA_RADS:
            targets[rad] = {}
            for day_type in ['weekday', 'thu', 'weekend']:
                if ira_total_availability_shares[day_type] > 0:
                    rad_share = ira_available_days[rad][day_type] / total_working_days_ytd[day_type]
                    targets[rad][day_type] = (total_working_days_ytd[day_type] / ira_total_availability_shares[day_type]) * rad_share
                else:
                    targets[rad][day_type] = 0.0
        
        # NO TARGETS FOR MRI SECTION - targets only calculated for GEN and IRA
        
        return targets
    
    def extract_month_year(self):
        """Extract month and year from filename"""
        import os
        import re
        
        filename = os.path.basename(self.excel_path)
        
        months = {
            'january': 1, 'february': 2, 'march': 3, 'april': 4,
            'may': 5, 'june': 6, 'july': 7, 'august': 8,
            'september': 9, 'october': 10, 'november': 11, 'december': 12
        }
        
        filename_lower = filename.lower()
        for month_name, month_num in months.items():
            if month_name in filename_lower:
                year_match = re.search(r'20\d{2}', filename)
                year = int(year_match.group()) if year_match else 2025
                return month_num, year
        
        return 11, 2025
    
    def cache_ytd_totals(self):
        """Cache YTD totals before making any assignments"""
        cache = {}
        all_rads = set(GEN_RADS + IRA_RADS + MRI_RADS)
        
        for rad in all_rads:
            if rad in ['TA', 'SG']:
                continue
            
            row = GEN_ROWS.get(rad) or IRA_ROWS.get(rad) or MRI_ROWS.get(rad)
            if not row:
                continue
            
            weekday_ytd = self.ws.cell(row, COL_AM).value
            thu_ytd = self.ws.cell(row, COL_AN).value
            weekend_ytd = self.ws.cell(row, COL_AO).value
            
            try:
                weekday_val = float(weekday_ytd) if weekday_ytd is not None else 0.0
            except (ValueError, TypeError):
                weekday_val = 0.0
            
            try:
                thu_val = float(thu_ytd) if thu_ytd is not None else 0.0
            except (ValueError, TypeError):
                thu_val = 0.0
            
            try:
                weekend_val = float(weekend_ytd) if weekend_ytd is not None else 0.0
            except (ValueError, TypeError):
                weekend_val = 0.0
            
            cache[rad] = {
                'weekday': weekday_val,
                'thu': thu_val,
                'weekend': weekend_val
            }
        
        return cache
    
    def identify_locked_assignments(self):
        """Identify cells that already have 'X' marks"""
        locked = defaultdict(dict)
        
        for rad, row in GEN_ROWS.items():
            if rad in ['TA']:
                continue
            for day in range(1, self.days_in_month + 1):
                col = day + 3
                cell = self.ws.cell(row, col)
                if cell.value and str(cell.value).strip().upper() == 'X':
                    locked['GEN'][day] = rad
        
        for rad, row in IRA_ROWS.items():
            for day in range(1, self.days_in_month + 1):
                col = day + 3
                cell = self.ws.cell(row, col)
                if cell.value and str(cell.value).strip().upper() == 'X':
                    locked['IRA'][day] = rad
        
        for rad, row in MRI_ROWS.items():
            for day in range(1, self.days_in_month + 1):
                col = day + 3
                cell = self.ws.cell(row, col)
                if cell.value and str(cell.value).strip().upper() == 'X':
                    locked['MRI'][day] = rad
        
        return locked
    
    def identify_vacation_days(self):
        """Identify vacation days (dark blue cells)"""
        vacations = defaultdict(set)
        
        for rad in set(GEN_RADS + IRA_RADS + MRI_RADS):
            if rad in ['TA', 'SG']:
                continue
            
            if rad in IRA_RADS:
                row = IRA_ROWS[rad]
            elif rad in GEN_ROWS:
                row = GEN_ROWS[rad]
            else:
                row = MRI_ROWS.get(rad)
            
            if not row:
                continue
            
            for day in range(1, self.days_in_month + 1):
                col = day + 3
                cell = self.ws.cell(row, col)
                
                if cell.fill and cell.fill.start_color:
                    color = cell.fill.start_color.rgb
                    if color and isinstance(color, str):
                        if color.upper() in ['FF4472C4', 'FF00B0F0', 'FF0070C0', 'FF002060']:
                            vacations[rad].add(day)
        
        return vacations
    
    def get_day_type(self, day):
        """Determine if day is weekday, Thursday, or weekend"""
        date = datetime(self.year, self.month, day)
        weekday = date.weekday()
        
        if weekday == 3:
            return 'thu'
        elif weekday in [4, 5]:
            return 'weekend'
        else:
            return 'weekday'
    
    def is_available(self, rad, day, section='GEN'):
        """Check if rad is available for assignment (hard constraints only)"""
        # Vacation check (hard constraint)
        if day in self.vacation_days.get(rad, set()):
            return False
        
        # Pre-vacation check (day before vacation) (hard constraint)
        if (day + 1) in self.vacation_days.get(rad, set()):
            return False
        
        # Special requests off (hard constraint)
        if rad in self.special_requests_off and day in self.special_requests_off[rad]:
            return False
        
        # Section eligibility
        if section == 'GEN' and rad not in GEN_RADS_WITH_IRA:
            return False
        if section == 'IRA' and rad not in IRA_RADS:
            return False
        if section == 'MRI' and rad not in MRI_RADS:
            return False
        
        # GEN weekend limit check (hard limit: 3 weekend days)
        if section == 'GEN':
            day_type = self.get_day_type(day)
            if day_type == 'weekend':
                gen_weekend_count = sum(1 for d, r in self.assignments['GEN'].items() 
                                       if r == rad and self.get_day_type(d) == 'weekend')
                if gen_weekend_count >= MAX_MONTHLY_WEEKENDS_GEN:
                    return False
        
        # Monthly limits per section
        if section == 'GEN' and self.gen_monthly_total[rad] >= MAX_MONTHLY_GEN:
            return False
        if section == 'IRA' and self.ira_monthly_total[rad] >= MAX_MONTHLY_IRA:
            return False
        if section == 'MRI' and self.mri_monthly_total[rad] >= MAX_MONTHLY_MRI:
            return False
        
        return True
    
    def calculate_workload_score(self, rad, day, day_type, section='GEN'):
        """Calculate workload score balancing toward YTD targets"""
        target = self.ytd_targets.get(rad, {'weekday': 0, 'thu': 0, 'weekend': 0})
        target_ytd = target[day_type]
        
        actual_ytd = self.ytd_cache[rad][day_type]
        monthly_total = self.monthly_counts[rad][day_type]
        projected_ytd = actual_ytd + monthly_total
        
        deviation_from_target = projected_ytd - target_ytd
        score = deviation_from_target * 10
        
        # Soft constraint penalty (discouraged but can assign)
        if rad in self.soft_constraints_off and day in self.soft_constraints_off[rad]:
            score += SOFT_CONSTRAINT_PENALTY
        
        # For GEN: heavily penalize IRA rads (IG, MF, AS) - use them only if necessary
        if section == 'GEN' and rad in IRA_RADS:
            score += 500  # Heavy penalty to deprioritize IRA rads for GEN
        
        # Weekend weighting (more burdensome)
        if day_type == 'weekend':
            score = score * 3
            
            # Penalize if already has 2+ weekend days
            if monthly_total >= 2:
                score += 50
        
        # Consecutive weekend penalty for GEN
        if section == 'GEN' and day_type == 'weekend':
            date = datetime(self.year, self.month, day)
            prev_weekend_days = []
            
            for days_back in range(5, 8):
                prev_date = date - timedelta(days=days_back)
                if prev_date.month == self.month:
                    prev_day = prev_date.day
                    if self.get_day_type(prev_day) == 'weekend':
                        prev_weekend_days.append(prev_day)
            
            for prev_day in prev_weekend_days:
                if self.assignments['GEN'].get(prev_day) == rad:
                    score += CONSECUTIVE_WEEKEND_PENALTY
        
        # Consecutive day penalty for GEN
        if section == 'GEN':
            if day > 1 and self.assignments['GEN'].get(day - 1) == rad:
                score += 200  # Heavy penalty for consecutive days
        
        return score
    
    def assign_gen_thursday_saturday(self):
        """Assign GEN Thursday-Saturday pairs"""
        print("\n=== Assigning GEN Thursday-Saturday pairs ===")
        
        for day in range(1, self.days_in_month + 1):
            if self.get_day_type(day) != 'thu':
                continue
            
            saturday_day = day + 2
            if saturday_day > self.days_in_month:
                continue
            
            date_check = datetime(self.year, self.month, saturday_day)
            if date_check.weekday() != 5:
                continue
            
            # Check for locked assignment
            if day in self.locked_assignments['GEN']:
                assigned_rad = self.locked_assignments['GEN'][day]
                self.assignments['GEN'][day] = assigned_rad
                self.assignments['GEN'][saturday_day] = assigned_rad
                self.monthly_counts[assigned_rad]['thu'] += 1
                self.monthly_counts[assigned_rad]['weekend'] += 1
                self.gen_monthly_total[assigned_rad] += 2
                print(f"  Day {day} (Thu) + {saturday_day} (Sat) -> {assigned_rad} [LOCKED]")
                continue
            
            # Find available rads - PREFER the 9 primary GEN rads
            primary_available = [r for r in GEN_RADS 
                                if self.is_available(r, day, 'GEN') 
                                and self.is_available(r, saturday_day, 'GEN')]
            
            # If no primary rads available, try IRA rads
            if not primary_available:
                available_rads = [r for r in GEN_RADS_WITH_IRA 
                                 if self.is_available(r, day, 'GEN') 
                                 and self.is_available(r, saturday_day, 'GEN')]
            else:
                available_rads = primary_available
            
            if not available_rads:
                print(f"  WARNING: No available rads for Thu {day} + Sat {saturday_day}")
                continue
            
            # Calculate scores
            scores = {}
            for rad in available_rads:
                thu_score = self.calculate_workload_score(rad, day, 'thu', 'GEN')
                sat_score = self.calculate_workload_score(rad, saturday_day, 'weekend', 'GEN')
                scores[rad] = thu_score + sat_score
            
            best_rad = min(scores, key=scores.get)
            self.assignments['GEN'][day] = best_rad
            self.assignments['GEN'][saturday_day] = best_rad
            self.monthly_counts[best_rad]['thu'] += 1
            self.monthly_counts[best_rad]['weekend'] += 1
            self.gen_monthly_total[best_rad] += 2
            print(f"  Day {day} (Thu) + {saturday_day} (Sat) -> {best_rad} (score: {scores[best_rad]:.1f})")
 
    def assign_ira_triplets(self):
        """Assign IRA Thursday-Friday-Saturday triplets ONLY"""
        print("\n=== Assigning IRA Thursday-Friday-Saturday triplets ===")
        
        # Track weekend assignments for balance (each rad should get ~1-2 weekends max)
        ira_weekend_count = {rad: 0 for rad in IRA_RADS}
        last_weekend_rad = None  # Track who had the previous weekend
        
        for day in range(1, self.days_in_month + 1):
            if self.get_day_type(day) != 'thu':
                continue
            
            fri_day = day + 1
            sat_day = day + 2
            
            if sat_day > self.days_in_month:
                continue
            
            # Check for locked assignment
            if day in self.locked_assignments['IRA']:
                assigned_rad = self.locked_assignments['IRA'][day]
                self.assignments['IRA'][day] = assigned_rad
                self.assignments['IRA'][fri_day] = assigned_rad
                self.assignments['IRA'][sat_day] = assigned_rad
                self.monthly_counts[assigned_rad]['thu'] += 1
                self.monthly_counts[assigned_rad]['weekend'] += 2
                self.ira_monthly_total[assigned_rad] += 3
                ira_weekend_count[assigned_rad] += 1
                last_weekend_rad = assigned_rad
                
                print(f"  Day {day}-{sat_day} (Thu-Fri-Sat) -> {assigned_rad} [LOCKED]")
                continue
            
            # Find available rads
            available_rads = [r for r in IRA_RADS 
                             if self.is_available(r, day, 'IRA') 
                             and self.is_available(r, fri_day, 'IRA')
                             and self.is_available(r, sat_day, 'IRA')]
            
            if not available_rads:
                print(f"  WARNING: No available IRA rads for {day}-{sat_day}")
                continue
            
            # Calculate scores with HEAVY emphasis on weekend balance AND avoiding consecutive weekends
            scores = {}
            for rad in available_rads:
                # Base YTD score
                thu_score = self.calculate_workload_score(rad, day, 'thu', 'IRA')
                fri_score = self.calculate_workload_score(rad, fri_day, 'weekend', 'IRA')
                sat_score = self.calculate_workload_score(rad, sat_day, 'weekend', 'IRA')
                base_score = thu_score + fri_score + sat_score
                
                # CRITICAL: Heavy penalty for consecutive weekends
                consecutive_penalty = 0
                if rad == last_weekend_rad:
                    consecutive_penalty = 1500  # Very high penalty for back-to-back weekends
                
                # CRITICAL: Add heavy penalty based on current weekend count
                # Goal: distribute weekends evenly (ideally 1-2 per rad, never 3+)
                weekend_penalty = ira_weekend_count[rad] * 500  # 500 points per weekend already assigned
                
                # Extra heavy penalty if rad already has 2+ weekends
                if ira_weekend_count[rad] >= 2:
                    weekend_penalty += 2000
                
                scores[rad] = base_score + consecutive_penalty + weekend_penalty
            
            best_rad = min(scores, key=scores.get)
            self.assignments['IRA'][day] = best_rad
            self.assignments['IRA'][fri_day] = best_rad
            self.assignments['IRA'][sat_day] = best_rad
            self.monthly_counts[best_rad]['thu'] += 1
            self.monthly_counts[best_rad]['weekend'] += 2
            self.ira_monthly_total[best_rad] += 3
            ira_weekend_count[best_rad] += 1
            last_weekend_rad = best_rad
            
            print(f"  Day {day}-{sat_day} (Thu-Fri-Sat) -> {best_rad} (score: {scores[best_rad]:.1f}, weekend #{ira_weekend_count[best_rad]})")
        
        # Print weekend distribution summary
        print("\n  IRA Weekend Distribution:")
        for rad in IRA_RADS:
            print(f"    {rad}: {ira_weekend_count[rad]} weekend(s)")

    def assign_remaining_days(self, section, eligible_rads):
        """Assign remaining unassigned days for a section"""
        print(f"\n=== Assigning remaining {section} days ===")
        
        # Count already assigned days
        assigned_count = len(self.assignments[section])
        print(f"Already assigned: {assigned_count} days")
        print(f"Need to assign: {self.days_in_month - assigned_count} more days")
        
        for day in range(1, self.days_in_month + 1):
            # Skip if already assigned
            if day in self.assignments[section]:
                continue
            
            # Check for locked assignment
            if day in self.locked_assignments[section]:
                assigned_rad = self.locked_assignments[section][day]
                self.assignments[section][day] = assigned_rad
                day_type = self.get_day_type(day)
                self.monthly_counts[assigned_rad][day_type] += 1
                
                if section == 'GEN':
                    self.gen_monthly_total[assigned_rad] += 1
                elif section == 'IRA':
                    self.ira_monthly_total[assigned_rad] += 1
                elif section == 'MRI':
                    self.mri_monthly_total[assigned_rad] += 1
                
                print(f"  Day {day} ({day_type}) -> {assigned_rad} [LOCKED]")
                continue
            
            day_type = self.get_day_type(day)
            
            # For GEN: prefer primary GEN rads, use IRA rads only if necessary
            if section == 'GEN':
                primary_available = [r for r in GEN_RADS if self.is_available(r, day, section)]
                if primary_available:
                    available_rads = primary_available
                else:
                    # Fall back to IRA rads if no primary rads available
                    available_rads = [r for r in eligible_rads if self.is_available(r, day, section)]
            else:
                available_rads = [r for r in eligible_rads if self.is_available(r, day, section)]
            
            if not available_rads:
                print(f"  ERROR: No available rads for {section} day {day}")
                # Try to find ANY rad by relaxing constraints
                print(f"    Attempting fallback assignment...")
                for rad in eligible_rads:
                    # Check only critical constraints (vacation, section eligibility)
                    if (day not in self.vacation_days.get(rad, set()) and
                        (day + 1) not in self.vacation_days.get(rad, set())):
                        
                        if section == 'GEN' and rad not in GEN_RADS_WITH_IRA:
                            continue
                        if section == 'IRA' and rad not in IRA_RADS:
                            continue
                        if section == 'MRI' and rad not in MRI_RADS:
                            continue
                        
                        available_rads = [rad]
                        print(f"    Fallback: Found {rad} by relaxing limits")
                        break
                
                if not available_rads:
                    print(f"  CRITICAL ERROR: Cannot assign {section} for day {day}")
                    continue
            
            # Calculate scores
            scores = {rad: self.calculate_workload_score(rad, day, day_type, section) 
                     for rad in available_rads}
            
            best_rad = min(scores, key=scores.get)
            self.assignments[section][day] = best_rad
            self.monthly_counts[best_rad][day_type] += 1
            
            if section == 'GEN':
                self.gen_monthly_total[best_rad] += 1
            elif section == 'IRA':
                self.ira_monthly_total[best_rad] += 1
            elif section == 'MRI':
                self.mri_monthly_total[best_rad] += 1
            
            print(f"  Day {day} ({day_type}) -> {best_rad} (score: {scores[best_rad]:.1f})")
        
        # Final verification
        final_count = len(self.assignments[section])
        print(f"\n{section} Assignment Summary: {final_count}/{self.days_in_month} days assigned")
        if final_count < self.days_in_month:
            missing_days = [d for d in range(1, self.days_in_month + 1) if d not in self.assignments[section]]
            print(f"  WARNING: Missing days: {missing_days}")

    def assign_ira_remaining_weekdays(self):
        """Assign remaining IRA weekdays (Sun-Wed) - can be split flexibly between IRA rads"""
        print("\n=== Assigning remaining IRA weekdays (Sun-Wed) ===")
        print("Note: Weekday assignments can be split flexibly to minimize 3-rad days")
        
        # Count already assigned IRA days
        assigned_count = len(self.assignments['IRA'])
        remaining_days = [d for d in range(1, self.days_in_month + 1) 
                         if d not in self.assignments['IRA']]
        
        print(f"Already assigned: {assigned_count} days (Thu-Fri-Sat triplets)")
        print(f"Need to assign: {len(remaining_days)} more weekdays")
        
        for day in remaining_days:
            # Check for locked assignment
            if day in self.locked_assignments['IRA']:
                assigned_rad = self.locked_assignments['IRA'][day]
                self.assignments['IRA'][day] = assigned_rad
                day_type = self.get_day_type(day)
                self.monthly_counts[assigned_rad][day_type] += 1
                self.ira_monthly_total[assigned_rad] += 1
                print(f"  Day {day} ({day_type}) -> {assigned_rad} [LOCKED]")
                continue
            
            day_type = self.get_day_type(day)
            
            # Find available IRA rads
            available_rads = [r for r in IRA_RADS if self.is_available(r, day, 'IRA')]
            
            if not available_rads:
                print(f"  ERROR: No available IRA rads for day {day}")
                continue
            
            # Calculate scores - PRIORITIZE IRA rads who can also do MRI to minimize 3-rad days
            scores = {}
            for rad in available_rads:
                base_score = self.calculate_workload_score(rad, day, day_type, 'IRA')
                
                # CRITICAL: Heavy bonus (negative score) if this IRA rad can also do MRI
                # This reduces 3-rad days by making IRA=MRI on 2-rad days
                mri_bonus = 0
                if rad in MRI_RADS:
                    mri_bonus = -200  # Strong incentive to use IRA rads who can do MRI
                
                scores[rad] = base_score + mri_bonus
            
            best_rad = min(scores, key=scores.get)
            self.assignments['IRA'][day] = best_rad
            self.monthly_counts[best_rad][day_type] += 1
            self.ira_monthly_total[best_rad] += 1
            
            mri_note = " (can do MRI - helps minimize 3-rad days)" if best_rad in MRI_RADS else ""
            print(f"  Day {day} ({day_type}) -> {best_rad} (score: {scores[best_rad]:.1f}){mri_note}")
        
        # Final verification
        final_count = len(self.assignments['IRA'])
        print(f"\nIRA Assignment Summary: {final_count}/{self.days_in_month} days assigned")
        
        # CRITICAL FIX: Ensure all days are covered (catch orphaned weekends like Day 1 Saturday)
        if final_count < self.days_in_month:
            missing_days = [d for d in range(1, self.days_in_month + 1) 
                           if d not in self.assignments['IRA']]
            print(f"  WARNING: {len(missing_days)} days still unassigned: {missing_days}")
            print(f"  Assigning missing days now...")
            
            for day in missing_days:
                # Check for locked assignment first
                if day in self.locked_assignments['IRA']:
                    assigned_rad = self.locked_assignments['IRA'][day]
                    self.assignments['IRA'][day] = assigned_rad
                    day_type = self.get_day_type(day)
                    self.monthly_counts[assigned_rad][day_type] += 1
                    self.ira_monthly_total[assigned_rad] += 1
                    print(f"    Day {day} -> {assigned_rad} [LOCKED]")
                    continue
                
                day_type = self.get_day_type(day)
                
                # Find available IRA rads
                available_rads = [r for r in IRA_RADS if self.is_available(r, day, 'IRA')]
                
                if not available_rads:
                    # If no one available due to constraints, use all IRA rads
                    print(f"    WARNING: No available rads for day {day}, using all IRA rads")
                    available_rads = IRA_RADS
                
                # Calculate scores
                scores = {}
                for rad in available_rads:
                    base_score = self.calculate_workload_score(rad, day, day_type, 'IRA')
                    mri_bonus = -200 if rad in MRI_RADS else 0
                    scores[rad] = base_score + mri_bonus
                
                best_rad = min(scores, key=scores.get)
                self.assignments['IRA'][day] = best_rad
                self.monthly_counts[best_rad][day_type] += 1
                self.ira_monthly_total[best_rad] += 1
                print(f"    Day {day} -> {best_rad} (score: {scores[best_rad]:.1f})")
            
            # Re-verify
            final_count = len(self.assignments['IRA'])
            print(f"\n  After fix: {final_count}/{self.days_in_month} days assigned")
            if final_count == self.days_in_month:
                print(f"  ✓ All IRA days now assigned!")


    def assign_mri_optimized(self):
        """Assign MRI minimizing 3-rad days, with fair distribution of MRI-only calls"""
        print("\n=== Assigning MRI (optimizing for 2-rad days) ===")
        
        # Track MRI-only assignments per GEN rad for fair distribution
        mri_only_assignments = defaultdict(int)
        
        for day in range(1, self.days_in_month + 1):
            # Skip if already assigned
            if day in self.assignments['MRI']:
                print(f"  Day {day} -> Already assigned to {self.assignments['MRI'][day]} [SKIPPING]")
                continue
            
            # Check for locked assignment
            if day in self.locked_assignments['MRI']:
                assigned_rad = self.locked_assignments['MRI'][day]
                self.assignments['MRI'][day] = assigned_rad
                day_type = self.get_day_type(day)
                self.monthly_counts[assigned_rad][day_type] += 1
                self.mri_monthly_total[assigned_rad] += 1
                print(f"  Day {day} ({day_type}) -> {assigned_rad} [LOCKED]")
                continue
            
            gen_rad = self.assignments['GEN'].get(day)
            ira_rad = self.assignments['IRA'].get(day)
            day_type = self.get_day_type(day)
            
            # Strategy 1: Same as GEN (if GEN != IRA and GEN can do MRI)
            if gen_rad and gen_rad != ira_rad and gen_rad in MRI_RADS:
                if self.is_available(gen_rad, day, 'MRI'):
                    self.assignments['MRI'][day] = gen_rad
                    self.monthly_counts[gen_rad][day_type] += 1
                    self.mri_monthly_total[gen_rad] += 1
                    print(f"  Day {day} ({day_type}) -> {gen_rad} (same as GEN = 2-rad day)")
                    continue
            
            # Strategy 2: Same as IRA (if IRA can do MRI)
            if ira_rad and ira_rad in MRI_RADS:
                if self.is_available(ira_rad, day, 'MRI'):
                    self.assignments['MRI'][day] = ira_rad
                    self.monthly_counts[ira_rad][day_type] += 1
                    self.mri_monthly_total[ira_rad] += 1
                    print(f"  Day {day} ({day_type}) -> {ira_rad} (same as IRA = 2-rad day)")
                    continue
            
            # Strategy 3: 3-rad day unavoidable - PREFER GEN rads who can do MRI
            print(f"  Day {day} ({day_type}) -> 3-rad day required")
            
            # First, try GEN rads who can do MRI (excluding current GEN/IRA rads)
            gen_mri_available = [r for r in GEN_RADS 
                                if r in MRI_RADS 
                                and r != gen_rad 
                                and r != ira_rad
                                and self.is_available(r, day, 'MRI')]
            
            if gen_mri_available:
                # Calculate scores with fair distribution emphasis
                scores = {}
                for rad in gen_mri_available:
                    base_score = self.calculate_workload_score(rad, day, day_type, 'MRI')
                    
                    # CRITICAL: Balance MRI-only assignments across GEN rads
                    # Heavy penalty based on how many MRI-only calls this rad already has
                    balance_penalty = mri_only_assignments[rad] * 400
                    
                    scores[rad] = base_score + balance_penalty
                
                best_rad = min(scores, key=scores.get)
                self.assignments['MRI'][day] = best_rad
                self.monthly_counts[best_rad][day_type] += 1
                self.mri_monthly_total[best_rad] += 1
                mri_only_assignments[best_rad] += 1
                
                print(f"    -> {best_rad} (GEN rad, MRI-only count: {mri_only_assignments[best_rad]}, score: {scores[best_rad]:.1f})")
                continue
            
            # Fallback: Try IRA rads who can do MRI (shouldn't happen often with optimized IRA weekday assignments)
            ira_mri_available = [r for r in IRA_RADS 
                                if r in MRI_RADS 
                                and r != gen_rad 
                                and r != ira_rad
                                and self.is_available(r, day, 'MRI')]
            
            if ira_mri_available:
                scores = {rad: self.calculate_workload_score(rad, day, day_type, 'MRI') 
                         for rad in ira_mri_available}
                best_rad = min(scores, key=scores.get)
                self.assignments['MRI'][day] = best_rad
                self.monthly_counts[best_rad][day_type] += 1
                self.mri_monthly_total[best_rad] += 1
                print(f"    -> {best_rad} (IRA rad, score: {scores[best_rad]:.1f})")
                continue
            
            # Last resort: Any available MRI rad
            available_rads = [r for r in MRI_RADS if self.is_available(r, day, 'MRI')]
            
            if not available_rads:
                print(f"  ERROR: No available MRI rads for day {day}")
                continue
            
            scores = {rad: self.calculate_workload_score(rad, day, day_type, 'MRI') 
                     for rad in available_rads}
            best_rad = min(scores, key=scores.get)
            self.assignments['MRI'][day] = best_rad
            self.monthly_counts[best_rad][day_type] += 1
            self.mri_monthly_total[best_rad] += 1
            print(f"    -> {best_rad} (fallback, score: {scores[best_rad]:.1f})")
        
        # Print MRI-only distribution among GEN rads
        print("\n  MRI-Only Distribution (3-rad days):")
        gen_mri_rads = [r for r in GEN_RADS if r in MRI_RADS]
        for rad in sorted(gen_mri_rads):
            count = mri_only_assignments.get(rad, 0)
            if count > 0:
                print(f"    {rad}: {count} MRI-only call(s)")
        
        # Check balance
        if mri_only_assignments:
            counts = list(mri_only_assignments.values())
            max_diff = max(counts) - min(counts)
            if max_diff <= 1:
                print(f"  ✓ Fair distribution: max difference = {max_diff} (≤1)")
            else:
                print(f"  ⚠ Unbalanced: max difference = {max_diff} (should be ≤1)")
        
        # Final verification
        final_count = len(self.assignments['MRI'])
        print(f"\nMRI Assignment Summary: {final_count}/{self.days_in_month} days assigned")    
    
    def write_schedule_to_excel(self):
        """Write all assignments to Excel"""
        print("\n=== Writing schedule to Excel ===")
        
        wb_write = openpyxl.load_workbook(self.excel_path)
        ws_write = wb_write['Sheet1']
        
        # Verify assignment counts before writing
        gen_days = len(self.assignments['GEN'])
        ira_days = len(self.assignments['IRA'])
        mri_days = len(self.assignments['MRI'])
        
        print(f"Assignment counts: GEN={gen_days}, IRA={ira_days}, MRI={mri_days}")
        print(f"Expected: {self.days_in_month} days each")
        
        if gen_days != self.days_in_month:
            print(f"  WARNING: GEN has {gen_days} assignments (expected {self.days_in_month})")
        if ira_days != self.days_in_month:
            print(f"  WARNING: IRA has {ira_days} assignments (expected {self.days_in_month})")
        if mri_days != self.days_in_month:
            print(f"  WARNING: MRI has {mri_days} assignments (expected {self.days_in_month})")
        
        # Clear all existing X marks for unlocked cells
        print("\n=== Clearing existing assignments ===")
        
        for day in range(1, self.days_in_month + 1):
            col = day + 3
            
            # Clear GEN section
            for rad, row in GEN_ROWS.items():
                if rad in ['TA']:
                    continue
                if not (day in self.locked_assignments['GEN'] and self.locked_assignments['GEN'][day] == rad):
                    ws_write.cell(row, col).value = None
            
            # Clear IRA section
            for rad, row in IRA_ROWS.items():
                if not (day in self.locked_assignments['IRA'] and self.locked_assignments['IRA'][day] == rad):
                    ws_write.cell(row, col).value = None
            
            # Clear MRI section
            for rad, row in MRI_ROWS.items():
                if not (day in self.locked_assignments['MRI'] and self.locked_assignments['MRI'][day] == rad):
                    ws_write.cell(row, col).value = None
        
        print("Sections cleared")
        
        # Write GEN assignments
        print("\n=== Writing GEN assignments ===")
        for day, rad in self.assignments['GEN'].items():
            if day > self.days_in_month:
                print(f"  ERROR: Skipping GEN day {day} (beyond month end)")
                continue
            row = GEN_ROWS[rad]
            col = day + 3
            ws_write.cell(row, col, 'X')
        print(f"Wrote {len(self.assignments['GEN'])} GEN assignments")
        
        # Write IRA assignments
        print("\n=== Writing IRA assignments ===")
        for day, rad in self.assignments['IRA'].items():
            if day > self.days_in_month:
                print(f"  ERROR: Skipping IRA day {day} (beyond month end)")
                continue
            row = IRA_ROWS[rad]
            col = day + 3
            ws_write.cell(row, col, 'X')
        print(f"Wrote {len(self.assignments['IRA'])} IRA assignments")
        
        # Write MRI assignments
        print("\n=== Writing MRI assignments ===")
        for day, rad in self.assignments['MRI'].items():
            if day > self.days_in_month:
                print(f"  ERROR: Skipping MRI day {day} (beyond month end)")
                continue
            row = MRI_ROWS[rad]
            col = day + 3
            ws_write.cell(row, col, 'X')
        print(f"Wrote {len(self.assignments['MRI'])} MRI assignments")
        
        # Write YTD targets to columns AQ(43), AR(44), AS(45)
        # ONLY for GEN and IRA sections (NOT MRI)
        print("\n=== Writing YTD Targets to Excel ===")
        print(f"Column AQ = {COL_AQ}, Column AR = {COL_AR}, Column AS = {COL_AS}")
        
        # GEN section targets - ONLY write to the 9 primary GEN rad rows (7-15)
        # DO NOT write to IG/MF/AS rows (17-19) - they only get IRA targets
        print("\nGEN Section Targets (9 rads):")
        for rad in GEN_RADS:
            if rad in self.ytd_targets:
                row = GEN_ROWS[rad]
                target = self.ytd_targets[rad]
                ws_write.cell(row, COL_AQ, round(target['weekday'], 1))
                ws_write.cell(row, COL_AR, round(target['thu'], 1))
                ws_write.cell(row, COL_AS, round(target['weekend'], 1))
                print(f"  {rad} → Row {row}, AQ={target['weekday']:.1f}, AR={target['thu']:.1f}, AS={target['weekend']:.1f}")
        
        # IRA section targets - write to IRA section rows (24-26)
        print("\nIRA Section Targets (3 rads):")
        for rad in IRA_RADS:
            if rad in self.ytd_targets:
                row = IRA_ROWS[rad]
                target = self.ytd_targets[rad]
                ws_write.cell(row, COL_AQ, round(target['weekday'], 1))
                ws_write.cell(row, COL_AR, round(target['thu'], 1))
                ws_write.cell(row, COL_AS, round(target['weekend'], 1))
                print(f"  {rad} → Row {row}, AQ={target['weekday']:.1f}, AR={target['thu']:.1f}, AS={target['weekend']:.1f}")
        
        # MRI section - NO TARGETS WRITTEN
        print("\nMRI section: No targets written (as expected)")
        print("Note: IG/MF/AS (rows 17-19) have NO GEN targets, only IRA targets")
        
        output_path = self.excel_path.replace('.xlsx', '_SCHEDULED.xlsx')
        wb_write.save(output_path)
        print(f"\nSchedule saved to: {output_path}")
        print(f"Column AP (42) remains BLANK")
        return output_path

    def print_summary(self):
        """Print comprehensive summary report"""
        print("\n" + "="*80)
        print("SCHEDULE SUMMARY REPORT")
        print("="*80)
        print(f"Month: {calendar.month_name[self.month]} {self.year}")
        print(f"Total days: {self.days_in_month}")
        
        # Count 2-rad vs 3-rad days and categorize MRI-only days
        two_rad_days = 0
        three_rad_days = 0
        mri_only_thu_sat = 0  # Thursday, Friday, Saturday
        mri_only_weekday = 0  # Sunday, Monday, Tuesday, Wednesday
        
        mri_only_thu_sat_details = []
        mri_only_weekday_details = []
        
        for day in range(1, self.days_in_month + 1):
            gen_rad = self.assignments['GEN'].get(day)
            ira_rad = self.assignments['IRA'].get(day)
            mri_rad = self.assignments['MRI'].get(day)
            
            unique_rads = len(set([gen_rad, ira_rad, mri_rad]) - {None})
            if unique_rads == 2:
                two_rad_days += 1
            elif unique_rads == 3:
                three_rad_days += 1
                
                # Categorize 3-rad days by day of week
                date = datetime(self.year, self.month, day)
                weekday = date.weekday()  # 0=Mon, 1=Tue, 2=Wed, 3=Thu, 4=Fri, 5=Sat, 6=Sun
                
                if weekday in [3, 4, 5]:  # Thu, Fri, Sat
                    mri_only_thu_sat += 1
                    day_name = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][weekday]
                    mri_only_thu_sat_details.append(f"Day {day} ({day_name}): {mri_rad}")
                else:  # Sun, Mon, Tue, Wed
                    mri_only_weekday += 1
                    day_name = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][weekday]
                    mri_only_weekday_details.append(f"Day {day} ({day_name}): {mri_rad}")
        
        print(f"\n2-Rad Days: {two_rad_days}")
        print(f"3-Rad Days: {three_rad_days}")
        if (two_rad_days + three_rad_days) > 0:
            print(f"Optimization: {two_rad_days / (two_rad_days + three_rad_days) * 100:.1f}% 2-rad days")
        
        # MRI-only breakdown
        print(f"\nMRI-Only Breakdown (3-rad days):")
        print(f"  Thu-Fri-Sat (Weekend): {mri_only_thu_sat} days (ideally 0)")
        if mri_only_thu_sat_details:
            for detail in mri_only_thu_sat_details:
                print(f"    - {detail}")
        
        print(f"  Sun-Mon-Tue-Wed (Weekday): {mri_only_weekday} days (minimize)")
        if mri_only_weekday_details:
            for detail in mri_only_weekday_details:
                print(f"    - {detail}")
        
        # Verification
        if mri_only_thu_sat + mri_only_weekday == three_rad_days:
            print(f"  ✓ Verification: {mri_only_thu_sat} + {mri_only_weekday} = {three_rad_days} (correct)")
        else:
            print(f"  ⚠ WARNING: {mri_only_thu_sat} + {mri_only_weekday} ≠ {three_rad_days} (mismatch!)")
        
        # Verify YTD Target totals
        print("\n" + "-"*80)
        print("YTD TARGET VERIFICATION")
        print("-"*80)
        
        for section_name, rads in [('GEN', GEN_RADS), ('IRA', IRA_RADS)]:
            print(f"\n{section_name} Section:")
            for day_type in ['weekday', 'thu', 'weekend']:
                sum_target = sum(self.ytd_targets[rad][day_type] for rad in rads if rad in self.ytd_targets)
                print(f"  {day_type.capitalize():8} - Total Target: {sum_target:.2f}")
        
        print("\n" + "-"*80)
        print("MONTHLY ASSIGNMENTS PER RAD (with YTD vs Target)")
        print("-"*80)
        
        # GEN Section Summary
        print("\n=== GEN SECTION (9 rads) ===")
        print(f"{'Rad':<6} {'WD':<4} {'Thu':<4} {'WE':<4} {'Tot':<5} {'YTD WD':<8} {'Tgt WD':<8} {'YTD Thu':<9} {'Tgt Thu':<9} {'YTD WE':<8} {'Tgt WE':<8} {'Status':<15}")
        print("-"*80)
        
        for rad in sorted(GEN_RADS):
            monthly = self.monthly_counts[rad]
            ytd = self.ytd_cache[rad]
            
            total_monthly = monthly['weekday'] + monthly['thu'] + monthly['weekend']
            target = self.ytd_targets.get(rad, {'weekday': 0, 'thu': 0, 'weekend': 0})
            
            proj_wd = ytd['weekday'] + monthly['weekday']
            proj_thu = ytd['thu'] + monthly['thu']
            proj_we = ytd['weekend'] + monthly['weekend']
            
            status = ""
            if abs(proj_we - target['weekend']) <= 1:
                status = "✓ Balanced"
            elif proj_we < target['weekend'] - 1:
                status = "↓ Behind"
            else:
                status = "↑ Ahead"
            
            print(f"{rad:<6} {monthly['weekday']:<4} {monthly['thu']:<4} {monthly['weekend']:<4} "
                  f"{total_monthly:<5} {proj_wd:<8.1f} {target['weekday']:<8.1f} "
                  f"{proj_thu:<9.1f} {target['thu']:<9.1f} {proj_we:<8.1f} {target['weekend']:<8.1f} {status:<15}")
        
        # IRA Section Summary
        print("\n=== IRA SECTION (3 rads) ===")
        print(f"{'Rad':<6} {'WD':<4} {'Thu':<4} {'WE':<4} {'Tot':<5} {'YTD WD':<8} {'Tgt WD':<8} {'YTD Thu':<9} {'Tgt Thu':<9} {'YTD WE':<8} {'Tgt WE':<8} {'Status':<15}")
        print("-"*80)
        
        for rad in sorted(IRA_RADS):
            monthly = self.monthly_counts[rad]
            ytd = self.ytd_cache[rad]
            
            total_monthly = monthly['weekday'] + monthly['thu'] + monthly['weekend']
            target = self.ytd_targets.get(rad, {'weekday': 0, 'thu': 0, 'weekend': 0})
            
            proj_wd = ytd['weekday'] + monthly['weekday']
            proj_thu = ytd['thu'] + monthly['thu']
            proj_we = ytd['weekend'] + monthly['weekend']
            
            status = ""
            if abs(proj_we - target['weekend']) <= 1:
                status = "✓ Balanced"
            elif proj_we < target['weekend'] - 1:
                status = "↓ Behind"
            else:
                status = "↑ Ahead"
            
            print(f"{rad:<6} {monthly['weekday']:<4} {monthly['thu']:<4} {monthly['weekend']:<4} "
                  f"{total_monthly:<5} {proj_wd:<8.1f} {target['weekday']:<8.1f} "
                  f"{proj_thu:<9.1f} {target['thu']:<9.1f} {proj_we:<8.1f} {target['weekend']:<8.1f} {status:<15}")
        
        # MRI Section Summary (no targets, just assignments)
        print("\n=== MRI SECTION (8 rads) ===")
        print(f"{'Rad':<6} {'WD':<4} {'Thu':<4} {'WE':<4} {'Tot':<5}")
        print("-"*80)
        
        for rad in sorted(MRI_RADS):
            monthly = self.monthly_counts[rad]
            total_monthly = monthly['weekday'] + monthly['thu'] + monthly['weekend']
            
            if total_monthly > 0:  # Only show rads with MRI assignments
                print(f"{rad:<6} {monthly['weekday']:<4} {monthly['thu']:<4} {monthly['weekend']:<4} {total_monthly:<5}")
        
        print("\n" + "-"*80)
        print("CONSTRAINT CHECK")
        print("-"*80)
        
        violations = []
        
        # GEN consecutive days check
        for day in range(1, self.days_in_month):
            if day in self.assignments['GEN'] and (day + 1) in self.assignments['GEN']:
                if self.assignments['GEN'][day] == self.assignments['GEN'][(day + 1)]:
                    violations.append(f"GEN consecutive days: {self.assignments['GEN'][day]} on days {day}-{day+1}")
        
        # GEN consecutive weekends check
        for day in range(1, self.days_in_month - 6):
            if self.get_day_type(day) == 'weekend':
                rad = self.assignments['GEN'].get(day)
                if rad:
                    for future_day in range(day + 5, min(day + 8, self.days_in_month + 1)):
                        if self.get_day_type(future_day) == 'weekend':
                            if self.assignments['GEN'].get(future_day) == rad:
                                violations.append(f"GEN consecutive weekends: {rad} on days {day} and {future_day}")
        
        # Vacation violations
        for rad, vacation_days in self.vacation_days.items():
            for day in vacation_days:
                if self.assignments['GEN'].get(day) == rad:
                    violations.append(f"Vacation violation: {rad} assigned GEN on vacation day {day}")
                if self.assignments['IRA'].get(day) == rad:
                    violations.append(f"Vacation violation: {rad} assigned IRA on vacation day {day}")
                if self.assignments['MRI'].get(day) == rad:
                    violations.append(f"Vacation violation: {rad} assigned MRI on vacation day {day}")
                
                if day > 1:
                    if self.assignments['GEN'].get(day - 1) == rad:
                        violations.append(f"Pre-vacation violation: {rad} assigned GEN day {day-1} before vacation")
                    if self.assignments['IRA'].get(day - 1) == rad:
                        violations.append(f"Pre-vacation violation: {rad} assigned IRA day {day-1} before vacation")
                    if self.assignments['MRI'].get(day - 1) == rad:
                        violations.append(f"Pre-vacation violation: {rad} assigned MRI day {day-1} before vacation")
        
        # GEN+IRA overload check
        for day in range(1, self.days_in_month + 1):
            gen_rad = self.assignments['GEN'].get(day)
            ira_rad = self.assignments['IRA'].get(day)
            if gen_rad and ira_rad and gen_rad == ira_rad:
                violations.append(f"Overload violation: {gen_rad} assigned both GEN and IRA on day {day}")
        
        # GEN Thursday-Saturday pairing check
        for day in range(1, self.days_in_month + 1):
            if self.get_day_type(day) == 'thu':
                saturday_day = day + 2
                if saturday_day <= self.days_in_month:
                    thu_rad = self.assignments['GEN'].get(day)
                    sat_rad = self.assignments['GEN'].get(saturday_day)
                    if thu_rad and sat_rad and thu_rad != sat_rad:
                        violations.append(f"GEN Thu-Sat pairing violation: {thu_rad} on Thu {day} but {sat_rad} on Sat {saturday_day}")
        
        # IRA triplet check
        for day in range(1, self.days_in_month + 1):
            if self.get_day_type(day) == 'thu':
                fri_day = day + 1
                sat_day = day + 2
                if sat_day <= self.days_in_month:
                    thu_rad = self.assignments['IRA'].get(day)
                    fri_rad = self.assignments['IRA'].get(fri_day)
                    sat_rad = self.assignments['IRA'].get(sat_day)
                    if thu_rad and fri_rad and thu_rad != fri_rad:
                        violations.append(f"IRA triplet violation: {thu_rad} on Thu {day} but {fri_rad} on Fri {fri_day}")
                    if thu_rad and sat_rad and thu_rad != sat_rad:
                        violations.append(f"IRA triplet violation: {thu_rad} on Thu {day} but {sat_rad} on Sat {sat_day}")
        
        # GEN weekend limit check
        for rad in GEN_RADS:
            weekend_count = sum(1 for day, r in self.assignments['GEN'].items() 
                               if r == rad and self.get_day_type(day) == 'weekend')
            if weekend_count > MAX_MONTHLY_WEEKENDS_GEN:
                violations.append(f"GEN weekend limit exceeded: {rad} has {weekend_count} weekend days (limit: {MAX_MONTHLY_WEEKENDS_GEN})")
        
        # Monthly limits check
        for rad in GEN_RADS_WITH_IRA:
            if self.gen_monthly_total[rad] > MAX_MONTHLY_GEN:
                violations.append(f"Monthly limit: {rad} has {self.gen_monthly_total[rad]} GEN days (limit: {MAX_MONTHLY_GEN})")
        
        for rad in IRA_RADS:
            if self.ira_monthly_total[rad] > MAX_MONTHLY_IRA:
                violations.append(f"Monthly limit: {rad} has {self.ira_monthly_total[rad]} IRA days (limit: {MAX_MONTHLY_IRA})")
        
        for rad in MRI_RADS:
            if self.mri_monthly_total[rad] > MAX_MONTHLY_MRI:
                violations.append(f"Monthly limit: {rad} has {self.mri_monthly_total[rad]} MRI days (limit: {MAX_MONTHLY_MRI})")
        
        # Coverage gap check
        for day in range(1, self.days_in_month + 1):
            if day not in self.assignments['GEN']:
                violations.append(f"Coverage gap: GEN not assigned for day {day}")
            if day not in self.assignments['IRA']:
                violations.append(f"Coverage gap: IRA not assigned for day {day}")
            if day not in self.assignments['MRI']:
                violations.append(f"Coverage gap: MRI not assigned for day {day}")
        
        if violations:
            print("⚠ VIOLATIONS FOUND:")
            for v in violations:
                print(f"  - {v}")
        else:
            print("✓ No constraint violations detected")
        
        print("\n" + "-"*80)
        print("WEEKEND DISTRIBUTION")
        print("-"*80)
        
        # GEN weekend distribution
        gen_weekend_counts = defaultdict(int)
        for day in range(1, self.days_in_month + 1):
            if self.get_day_type(day) == 'weekend':
                rad = self.assignments['GEN'].get(day)
                if rad:
                    gen_weekend_counts[rad] += 1
        
        total_weekends = sum(1 for day in range(1, self.days_in_month + 1) 
                            if self.get_day_type(day) == 'weekend')
        
        print(f"GEN Weekend Assignments (limit: {MAX_MONTHLY_WEEKENDS_GEN} days, {total_weekends} total weekend days):")
        for rad in sorted(GEN_RADS):
            count = gen_weekend_counts.get(rad, 0)
            status = "✓" if count <= MAX_MONTHLY_WEEKENDS_GEN else "⚠"
            print(f"  {status} {rad}: {count} weekend days")
        
        # IRA weekend distribution
        ira_weekend_counts = defaultdict(int)
        for day in range(1, self.days_in_month + 1):
            if self.get_day_type(day) == 'weekend':
                rad = self.assignments['IRA'].get(day)
                if rad:
                    ira_weekend_counts[rad] += 1
        
        print(f"\nIRA Weekend Assignments (no limit):")
        for rad in sorted(IRA_RADS):
            count = ira_weekend_counts.get(rad, 0)
            print(f"  {rad}: {count} weekend days")
        
        # MRI weekend distribution
        mri_weekend_counts = defaultdict(int)
        for day in range(1, self.days_in_month + 1):
            if self.get_day_type(day) == 'weekend':
                rad = self.assignments['MRI'].get(day)
                if rad:
                    mri_weekend_counts[rad] += 1
        
        print(f"\nMRI Weekend Assignments (no limit):")
        for rad in sorted(MRI_RADS):
            count = mri_weekend_counts.get(rad, 0)
            if count > 0:
                print(f"  {rad}: {count} weekend days")

    def generate_schedule(self):
        """Main method to generate the complete schedule"""
        print("="*80)
        print("ONCALL SCHEDULE GENERATOR")
        print("="*80)
        print(f"Processing: {self.excel_path}")
        print(f"Month: {calendar.month_name[self.month]} {self.year}")
        print(f"Days in month: {self.days_in_month}")
        
        # GET USER PREFERENCES FIRST
        self.get_user_preferences()
        
        print("\n" + "-"*80)
        print("CACHED YTD TOTALS & CALCULATED TARGETS")
        print("-"*80)
        print(f"{'Rad':<6} {'Start':<8} {'YTD WD':<8} {'YTD Thu':<8} {'YTD WE':<8} {'Tgt WD':<8} {'Tgt Thu':<8} {'Tgt WE':<8}")
        print("-"*80)
        
        for rad in sorted(self.ytd_cache.keys()):
            ytd = self.ytd_cache[rad]
            target = self.ytd_targets.get(rad, {'weekday': 0, 'thu': 0, 'weekend': 0})
            
            start_month, start_day = START_DATES.get(rad, (1, 1))
            start_date = f"{start_month}/{start_day}"
            
            print(f"{rad:<6} {start_date:<8} {ytd['weekday']:<8.1f} {ytd['thu']:<8.1f} {ytd['weekend']:<8.1f} "
                  f"{target['weekday']:<8.1f} {target['thu']:<8.1f} {target['weekend']:<8.1f}")
        
        print("\nNote: Targets based on available days through end of current month")
        print("      Sum of all rad targets = Sum of all rad actuals (per day type)")
        
        print("\n" + "-"*80)
        print("VACATION DAYS IDENTIFIED")
        print("-"*80)
        for rad in sorted(self.vacation_days.keys()):
            if self.vacation_days[rad]:
                days_str = ', '.join(map(str, sorted(self.vacation_days[rad])))
                print(f"{rad}: {days_str}")
        
        print("\n" + "-"*80)
        print("LOCKED ASSIGNMENTS (ON requests/holidays)")
        print("-"*80)
        for section in ['GEN', 'IRA', 'MRI']:
            if self.locked_assignments[section]:
                print(f"{section}:")
                for day, rad in sorted(self.locked_assignments[section].items()):
                    print(f"  Day {day}: {rad}")
            else:
                print(f"{section}: None")
        
        # UPDATED ASSIGNMENT ORDER
        self.assign_gen_thursday_saturday()          # 1. GEN Thu+Sat pairs first
        self.assign_ira_triplets()                   # 2. IRA Thu-Fri-Sat triplets
        self.assign_remaining_days('GEN', GEN_RADS_WITH_IRA)  # 3. Remaining GEN days
        self.assign_ira_remaining_weekdays()         # 4. Remaining IRA weekdays (flexible)
        self.assign_mri_optimized()                  # 5. MRI with 3-rad day minimization
        
        self.print_summary()
        
        output_path = self.write_schedule_to_excel()
        
        return output_path

def main():
    """Main entry point"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python oncall_scheduler.py <excel_file_path>")
        print("\nExample: python oncall_scheduler.py OnCall_Schedule_November_2025_BLANK.xlsx")
        return
    
    excel_path = sys.argv[1]
    
    scheduler = OnCallScheduler(excel_path)
    
    output_path = scheduler.generate_schedule()
    
    print("\n" + "="*80)
    print("SCHEDULE GENERATION COMPLETE")
    print("="*80)
    print(f"Output file: {output_path}")
    print("\nNext steps:")
    print("1. Review the schedule in Excel")
    print("2. Check for any constraint violations")
    print("3. Make manual adjustments if needed")
    print("\nNote: YTD totals auto-update via Excel formulas")


if __name__ == "__main__":
    main()
    
    
