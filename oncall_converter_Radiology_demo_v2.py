import openpyxl
import csv
import calendar
import sys
from datetime import datetime, timedelta
from pathlib import Path 

# Employee mapping dictionary - Modified structure
EMPLOYEE_MAP = {
    'allwo0f': {'emp_initials': 'AK', 'emp_roles': ['1056'], 'emp_name': 'Dr. Allison Livingston'},
    'audr95t': {'emp_initials': 'AO', 'emp_roles': ['1056'], 'emp_name': 'Dr. Audrey Randy'},
    'ellias4': {'emp_initials': 'AS', 'emp_roles': ['1056'], 'emp_name': 'Dr. Ankur Simran Ellison'},
    'lotta3': {'emp_initials': 'AT', 'emp_roles': ['1056'], 'emp_name': 'Dr. Angela Lotti'},
    'figeftr': {'emp_initials': 'FT', 'emp_roles': ['1056'], 'emp_name': 'Dr. Fernando Figer'},
    'hauser4': {'emp_initials': 'IG', 'emp_roles': ['1056'], 'emp_name': 'Dr. Irvin Garrett Hauser'},
    'kaisbam': {'emp_initials': 'LK', 'emp_roles': ['1056'], 'emp_name': 'Dr. Barry Midland Kaiser'},
    'bellam5': {'emp_initials': 'MB', 'emp_roles': ['1056'], 'emp_name': 'Dr. Monica Bella'},
    'chengme': {'emp_initials': 'MC', 'emp_roles': ['1056'], 'emp_name': 'Dr. Milkha Chengi'},
    'fakma0e': {'emp_initials': 'MF', 'emp_roles': ['1056'], 'emp_name': 'Dr. Maria Nargis'},
    'mumir4': {'emp_initials': 'MM', 'emp_roles': ['1056'], 'emp_name': 'Dr. Mir Miranda'},
    'nilanin': {'emp_initials': 'NN', 'emp_roles': ['1056'], 'emp_name': 'Dr. Nayan Nilani'},
    'hernapat': {'emp_initials': 'PR', 'emp_roles': ['1056'], 'emp_name': 'Dr. Paul Hernandez'},
    'gonzsa2': {'emp_initials': 'SG', 'emp_roles': ['1056'], 'emp_name': 'Dr. Gonzales, Salem'},
    'alitar3b': {'emp_initials': 'TA', 'emp_roles': ['1056'], 'emp_name': 'Dr. Tarzan Ali'},
    'ignaro5w': {'emp_initials': 'RI', 'emp_roles': ['1056'], 'emp_name': 'Dr. Roberta Ignatius'},
    '9999': {'emp_initials': 'TELE', 'emp_roles': ['58'], 'emp_name': 'Dr. Tele Radiology'}
}

# Create reverse lookup dictionary: initials -> employee number
INITIALS_TO_EMPNUM = {data['emp_initials']: emp_num for emp_num, data in EMPLOYEE_MAP.items()}

# Team configurations
TEAMS = {
    'Gen_CT': {'team_id': '114', 'work_cols': ['H', 'I'], 'oncall_rows': (5, 21)},
    'IRA': {'team_id': '115', 'work_cols': ['M'], 'oncall_rows': (24, 27)},
    'MRI': {'team_id': '116', 'work_cols': ['C'], 'oncall_rows': (30, 38)},
    'US': {'team_id': '126', 'work_cols': ['E'], 'oncall_rows': (5, 21)},
    'Fluoro': {'team_id': '127', 'work_cols': ['O'], 'oncall_rows': (5, 21)}
}

def col_letter_to_index(col_letter):
    """Convert column letter to index (A=1, B=2, etc.)"""
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def is_weekday(date):
    """Check if date is a weekday (Sun-Thu, 0=Monday, 6=Sunday)"""
    day_of_week = date.weekday()
    # In Python: Monday=0, Tuesday=1, Wednesday=2, Thursday=3, Friday=4, Saturday=5, Sunday=6
    # Weekdays are: Sunday(6), Monday(0), Tuesday(1), Wednesday(2), Thursday(3)
    # Weekends are: Friday(4), Saturday(5)
    return day_of_week in [6, 0, 1, 2, 3]  # Sun, Mon, Tue, Wed, Thu

def extract_month_year_from_filename(filename):
    """Extract month and year from filename"""
    import re
    
    filename_lower = filename.lower()
    
    # Try to find month name in filename
    month_num = None
    for month_idx in range(1, 13):
        month_name = calendar.month_name[month_idx].lower()
        month_abbr = calendar.month_abbr[month_idx].lower()
        if month_name in filename_lower or month_abbr in filename_lower:
            month_num = month_idx
            break
    
    # Try to find year (4-digit number)
    year_match = re.search(r'20\d{2}', filename)
    year = int(year_match.group()) if year_match else datetime.now().year
    
    return month_num, year

def get_employee_from_work_schedule(ws_work, day_num, col_letter):
    """Get employee initials from work schedule for a specific day and column""" 
    # NOTE: The function assumes that the Work Schedule cells contain employee initials (like "AK", "AT", "MC", etc.) that directly match the keys in INITIALS_TO_EMPNUM, which is exactly what the Excel files have!
    # Work schedule rows: 5:9, 13:17, 21:25, 29:33, 37:41
    row_ranges = [(5, 9), (13, 17), (21, 25), (29, 33), (37, 41)]
    
    col_idx = col_letter_to_index(col_letter)
    
    # Find the row corresponding to this day
    for row_range in row_ranges:
        for row in range(row_range[0], row_range[1] + 1):
            # Check column A for the day number
            day_cell = ws_work.cell(row=row, column=1).value
            
            # Handle date objects and other formats
            if day_cell:
                # If it's a datetime object, extract the day
                if isinstance(day_cell, datetime):
                    cell_day = day_cell.day
                else:
                    # Try to convert to string and extract day number
                    day_str = str(day_cell).strip()
                    # Try to extract day number if it's in format like "2-Nov" or just "2"
                    if '-' in day_str:
                        day_str = day_str.split('-')[0]
                    try:
                        cell_day = int(day_str)
                    except ValueError:
                        continue
                
                if cell_day == day_num:
                    # Get the employee initials from the specified column
                    emp_cell = ws_work.cell(row=row, column=col_idx).value
                    if emp_cell:
                        emp_str = str(emp_cell).strip().upper()
                        
                        # Handle combined readers (e.g., "AS/TELE", "MF/MM", "AK/TELE/MC")
                        if '/' in emp_str:
                            # Split by slash and filter out whitespace
                            readers = [r.strip() for r in emp_str.split('/') if r.strip()]
                            
                            # Try to find first non-TELE reader that exists in INITIALS_TO_EMPNUM
                            for reader in readers:
                                if reader != 'TELE' and reader in INITIALS_TO_EMPNUM:
                                    return reader
                            
                            # If all readers are TELE or no valid reader found, check if TELE exists
                            if 'TELE' in readers and 'TELE' in INITIALS_TO_EMPNUM:
                                return 'TELE'
                        else:
                            # Single reader - directly check if it's a valid key in INITIALS_TO_EMPNUM
                            if emp_str in INITIALS_TO_EMPNUM:
                                return emp_str
    
    return None

def get_employee_from_oncall_schedule(ws_oncall, day_num, row_start, row_end):
    """Get employee marked with X from oncall schedule for a specific day"""
    # Day columns start at D (column 4), so day 1 is column 4, day 2 is column 5, etc.
    day_col = 3 + day_num
    
    for row in range(row_start, row_end + 1):
        # Skip rows 23 and 29 as they are headers
        if row in [23, 29]:
            continue
        
        # Skip empty rows
        name_cell = ws_oncall.cell(row=row, column=1).value
        if not name_cell or str(name_cell).strip() == '':
            continue
            
        cell_value = ws_oncall.cell(row=row, column=day_col).value
        if cell_value:
            cell_str = str(cell_value).strip().upper()
            # Check for 'X' (and ignore '0' or empty)
            if cell_str == 'X':
                # Get employee name from column A, extract initials
                full_name = str(name_cell).strip().upper()
                
                # If name contains comma (LASTNAME, FIRSTNAME format), parse it
                if ',' in full_name:
                    parts = full_name.split(',')
                    last_name = parts[0].strip()
                    first_name = parts[1].strip() if len(parts) > 1 else ''
                    
                    # Try to match by last name (most reliable)
                    for emp_num, emp_data in EMPLOYEE_MAP.items():
                        emp_name_upper = emp_data['emp_name'].upper()
                        # Check if last name is in the employee name
                        if last_name in emp_name_upper:
                            # If we have first name, verify it also matches
                            if first_name:
                                if first_name in emp_name_upper:
                                    return emp_data['emp_initials']
                            else:
                                # No first name to check, last name match is sufficient
                                return emp_data['emp_initials']
                    
                    # If last name matching didn't work, try first name only
                    if first_name:
                        for emp_num, emp_data in EMPLOYEE_MAP.items():
                            if first_name in emp_data['emp_name'].upper():
                                return emp_data['emp_initials']
                else:
                    # No comma - try direct matching
                    for emp_num, emp_data in EMPLOYEE_MAP.items():
                        if full_name in emp_data['emp_name'].upper():
                            return emp_data['emp_initials']
    
    return None

def create_schedule_entry(emp_initials, team_id, start_date, start_time, end_date, end_time, date_format='%m/%d/%Y'):
    """Create a schedule entry dictionary"""
    # Convert initials to employee number
    if emp_initials not in INITIALS_TO_EMPNUM:
        return None
    
    emp_num = INITIALS_TO_EMPNUM[emp_initials]
    emp_data = EMPLOYEE_MAP[emp_num]
    
    return {
        'EMPLOYEE': emp_num,  # Now using the employee number (key) directly
        'TEAM': team_id,
        'STARTDATE': start_date.strftime(date_format),
        'STARTTIME': start_time,
        'ENDDATE': end_date.strftime(date_format),
        'ENDTIME': end_time,
        'ROLE': emp_data['emp_roles'][0],  # Get first role from the list
        'NOTES': '',
        'ORDER': '',
        'TEAMCOMMENT': ''
    }

def process_schedules(ws_work, ws_oncall, year, month):
    """Process both schedules and create output data"""
    output_data = []
    days_in_month = calendar.monthrange(year, month)[1]
    
    # Process each day of the month
    for day in range(1, days_in_month + 1):
        current_date = datetime(year, month, day)
        next_date = current_date + timedelta(days=1)
        
        is_weekday_flag = is_weekday(current_date)
        
        # Process each team
        for team_name, team_config in TEAMS.items():
            team_id = team_config['team_id']
            work_cols = team_config['work_cols']
            oncall_row_start, oncall_row_end = team_config['oncall_rows']
            
            if is_weekday_flag:
                # Weekdays (Sun-Thu): Process work schedule + oncall
                
                if team_name == 'Gen_CT':
                    # Gen_CT has 3 blocks on weekdays: 0700-1100, 1100-1530, 1530-0700
                    
                    # Block 1: 0700-1100 (Column H)
                    emp1 = get_employee_from_work_schedule(ws_work, day, work_cols[0])
                    if emp1:
                        entry = create_schedule_entry(emp1, team_id, current_date, '700', current_date, '1100')
                        if entry:
                            output_data.append(entry)
                    
                    # Block 2: 1100-1530 (Column I)
                    emp2 = get_employee_from_work_schedule(ws_work, day, work_cols[1])
                    if emp2:
                        entry = create_schedule_entry(emp2, team_id, current_date, '1100', current_date, '1530')
                        if entry:
                            output_data.append(entry)
                    
                    # Block 3: 1530-0700 next day (OnCall)
                    emp3 = get_employee_from_oncall_schedule(ws_oncall, day, oncall_row_start, oncall_row_end)
                    if emp3:
                        entry = create_schedule_entry(emp3, team_id, current_date, '1530', next_date, '700')
                        if entry:
                            output_data.append(entry)
                
                else:
                    # Other teams have 2 blocks on weekdays: 0700-1530, 1530-0700
                    
                    # Block 1: 0700-1530 (Work schedule)
                    emp1 = get_employee_from_work_schedule(ws_work, day, work_cols[0])
                    if emp1:
                        entry = create_schedule_entry(emp1, team_id, current_date, '700', current_date, '1530')
                        if entry:
                            output_data.append(entry)
                    
                    # Block 2: 1530-0700 next day (OnCall)
                    emp2 = get_employee_from_oncall_schedule(ws_oncall, day, oncall_row_start, oncall_row_end)
                    if emp2:
                        entry = create_schedule_entry(emp2, team_id, current_date, '1530', next_date, '700')
                        if entry:
                            output_data.append(entry)
            
            else:
                # Weekends (Fri-Sat): Only oncall, full 24 hours 0700-0700
                emp = get_employee_from_oncall_schedule(ws_oncall, day, oncall_row_start, oncall_row_end)
                if emp:
                    entry = create_schedule_entry(emp, team_id, current_date, '700', next_date, '700')
                    if entry:
                        output_data.append(entry)
    
    return output_data

def write_output_files(output_data, output_dir, filename_prefix):
    """Write output to both CSV and Excel formats"""
    
    # Write CSV file with caret delimiter
    csv_filename = output_dir / f"{filename_prefix}.csv"
    with open(csv_filename, 'w', newline='') as csvfile:
        fieldnames = ['EMPLOYEE', 'TEAM', 'STARTDATE', 'STARTTIME', 
                     'ENDDATE', 'ENDTIME', 'ROLE', 'NOTES', 'ORDER', 'TEAMCOMMENT']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter='^')
        writer.writeheader()
        writer.writerows(output_data)
    
    print(f"Created CSV: {csv_filename}")
    
    # Write Excel file
    xlsx_filename = output_dir / f"{filename_prefix}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Write headers
    headers = ['EMPLOYEE', 'TEAM', 'STARTDATE', 'STARTTIME', 
               'ENDDATE', 'ENDTIME', 'ROLE', 'NOTES', 'ORDER', 'TEAMCOMMENT']
    ws.append(headers)
    
    # Write data rows
    for row in output_data:
        ws.append([row[h] for h in headers])
    
    wb.save(xlsx_filename)
    print(f"Created Excel: {xlsx_filename}")

def main():
    # Check if command line arguments were provided
    if len(sys.argv) >= 3:
        work_schedule_path = sys.argv[1]
        oncall_schedule_path = sys.argv[2]
    else:
        # Prompt user for file paths
        print("OnCall Schedule Converter")
        print("=" * 50)
        work_schedule_path = input("Enter the path to Work Schedule file (.xlsx): ").strip()
        oncall_schedule_path = input("Enter the path to OnCall Schedule file (.xlsx): ").strip()
        
        # Remove quotes if user copied path with quotes
        work_schedule_path = work_schedule_path.strip('"').strip("'")
        oncall_schedule_path = oncall_schedule_path.strip('"').strip("'")
    
    work_file = Path(work_schedule_path)
    oncall_file = Path(oncall_schedule_path)
    
    if not work_file.exists():
        print(f"Error: Work schedule file not found at {work_schedule_path}")
        return
    
    if not oncall_file.exists():
        print(f"Error: OnCall schedule file not found at {oncall_schedule_path}")
        return
    
    # Output directory is the same as input files directory
    output_dir = work_file.parent
    
    # Load workbooks
    print(f"Loading Work Schedule: {work_file}")
    wb_work = openpyxl.load_workbook(work_file)
    ws_work = wb_work['WORK SCHEDULE']
    
    print(f"Loading OnCall Schedule: {oncall_file}")
    # data_only=True reads calculated values instead of formulas
    wb_oncall = openpyxl.load_workbook(oncall_file, data_only=True)
    ws_oncall = wb_oncall['Sheet1']
    
    # Extract month and year from filename
    filename = oncall_file.stem
    current_month, current_year = extract_month_year_from_filename(filename)
    
    if current_month is None:
        current_month = datetime.now().month
        print(f"Warning: Could not detect month from filename, using current month")
    
    print(f"\nProcessing schedules for {calendar.month_name[current_month]} {current_year}")
    
    # Process schedules
    output_data = process_schedules(ws_work, ws_oncall, current_year, current_month)
    
    print(f"\nGenerated {len(output_data)} schedule entries")
    
    # Write output files
    write_output_files(output_data, output_dir, "RadCall_import")
    
    print(f"\n✓ All files created successfully in: {output_dir}")

if __name__ == "__main__":
    main()
