import streamlit as st
import sys
import os
from pathlib import Path
import tempfile
import io
import csv
import calendar
from datetime import datetime
import openpyxl

# Import the converter module
import oncall_converter_Radiology_demo_v2 as rad_converter

# Configure page
st.set_page_config(
    page_title="Radiology - Schedule Management",
    page_icon="ü©ª",
    layout="centered",
    initial_sidebar_state="expanded"
)

# Sidebar with About section ONLY (no custom navigation)
with st.sidebar:
    #st.markdown("### About")
    st.info("**Project by JA RAD**")

# Main page title
st.title("ü©ª Radiology Schedule Management")
st.markdown("---")

# Section 1: Create Rad Work Schedule (Placeholder)
st.header("üìù Create Rad Work Schedule")
st.markdown("Upload a blank or partially filled Work Schedule template to generate a completed schedule.")
st.warning("‚ö†Ô∏è **In Development** - This feature is currently under development.")

work_template = st.file_uploader(
    #"Upload Work Schedule Template (Excel)", 
    type=['xlsx'],
    key='work_template',
    help="Upload the blank Work Schedule template Excel file"
)

if work_template:
    st.info("üìã Work Schedule creation feature coming soon...")

st.markdown("---")

# Section 2: Create Rad On-Call Schedule
st.header("üìÖ Create Rad On-Call Schedule")
st.markdown("Upload a blank or partially filled On-Call Schedule template to generate an optimized schedule.")

oncall_template = st.file_uploader(
    #"Upload On-Call Schedule Template (Excel)", 
    type=['xlsx'],
    key='oncall_template',
    help="Upload the blank On-Call schedule template Excel file"
)

if oncall_template:
    st.success("‚úÖ Template uploaded successfully")
    
    # Create expandable section for preferences
    with st.expander("‚öôÔ∏è Scheduling Preferences (Optional)", expanded=False):
        st.markdown("### Prior Month Information")
        st.markdown("Provide information about on-call assignments from the previous month to optimize scheduling.")
        
        # Section 1: GEN rads on last weekends of prior month
        st.markdown("#### 1. GEN Rads on Last Weekend of Prior Month")
        st.caption("These rads will NOT be assigned the first weekend of current month")
        
        col1, col2 = st.columns(2)
        with col1:
            gen_last_weekend_1 = st.selectbox(
                "GEN Rad #1",
                [""] + ['NN', 'MB', 'LK', 'PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'IG', 'MF', 'AS'],
                key='gen_last_weekend_1'
            )
        with col2:
            gen_last_weekend_2 = st.selectbox(
                "GEN Rad #2",
                [""] + ['NN', 'MB', 'LK', 'PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'IG', 'MF', 'AS'],
                key='gen_last_weekend_2'
            )
        
        # Section 2: GEN rad on last day of prior month
        st.markdown("#### 2. GEN Rad on Last Day of Prior Month")
        st.caption("This rad will NOT be assigned on days 1-2 of current month")
        
        gen_last_day = st.selectbox(
            "GEN Rad",
            [""] + ['NN', 'MB', 'LK', 'PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'IG', 'MF', 'AS'],
            key='gen_last_day'
        )
        
        # Section 3: IRA rad on last weekend of prior month
        st.markdown("#### 3. IRA Rad on Last Weekend of Prior Month")
        st.caption("This rad is discouraged but can be assigned during first week")
        
        ira_last_weekend = st.selectbox(
            "IRA Rad",
            [""] + ['IG', 'MF', 'AS'],
            key='ira_last_weekend'
        )
        
        # Section 4: Additional requests
        st.markdown("#### 4. Additional Preference Requests (OFF Days)")
        st.caption("‚ö†Ô∏è Add custom OFF requests (days rads should NOT work)")
        st.caption("üìå Note: Pre-filled X marks in your Excel template are automatically preserved as locked ON assignments")
        
        if 'additional_requests' not in st.session_state:
            st.session_state.additional_requests = []
        
        with st.form("add_request_form"):
            col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
            
            with col1:
                request_section = st.selectbox(
                    "Section",
                    ["GEN", "IRA", "MRI"],
                    key='req_section'
                )
            
            with col2:
                if request_section == "GEN":
                    rad_options = [""] + ['NN', 'MB', 'LK', 'PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'IG', 'MF', 'AS']
                elif request_section == "IRA":
                    rad_options = [""] + ['IG', 'MF', 'AS']
                else:  # MRI
                    rad_options = [""] + ['PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'MF', 'AS']
                
                request_rad = st.selectbox("Rad", rad_options, key='req_rad')
            
            with col3:
                request_day = st.number_input("Day", min_value=1, max_value=31, value=1, key='req_day')
            
            with col4:
                st.write("")  # Spacer
                st.write("")  # Spacer
                request_hard = st.checkbox("Hard", key='req_hard', help="Hard constraint = cannot assign")
            
            add_button = st.form_submit_button("‚ûï Add Request")
            
            if add_button:
                if not request_rad or request_rad == "":
                    st.warning("‚ö†Ô∏è Please select a Rad before adding request")
                else:
                    constraint_type = "HARD" if request_hard else "SOFT"
                    request_str = f"{request_section}/{request_rad}/Day{request_day}/OFF ({constraint_type})"
                    st.session_state.additional_requests.append({
                        'section': request_section,
                        'rad': request_rad,
                        'day': request_day,
                        'hard': request_hard,
                        'display': request_str
                    })
                    st.rerun()
        
        # Display added requests
        if st.session_state.additional_requests:
            st.markdown("**Added Requests:**")
            for idx, req in enumerate(st.session_state.additional_requests):
                col1, col2 = st.columns([4, 1])
                with col1:
                    st.caption(f"‚Ä¢ {req['display']}")
                with col2:
                    if st.button("üóëÔ∏è", key=f"del_{idx}"):
                        st.session_state.additional_requests.pop(idx)
                        st.rerun()
    
    st.markdown("---")
    
    # Generate button - MINIMAL VERSION - Just make it work!
    if st.button("üöÄ Generate Schedule", type="primary", use_container_width=True):
        try:
            with st.spinner("Generating optimized schedule..."):
                # Import the scheduler
                import create_oncall_schedule_v3 as scheduler_module
                from collections import defaultdict
                from datetime import timedelta
                
                # Save uploaded file temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(oncall_template.getvalue())
                    temp_path = tmp.name
                
                # Create scheduler instance
                scheduler = scheduler_module.OnCallScheduler(temp_path)
                
                # Apply user preferences programmatically
                scheduler.special_requests_off = defaultdict(set)
                scheduler.soft_constraints_off = defaultdict(set)
                
                # Section 1: GEN last weekend
                last_weekend_gen = [r for r in [gen_last_weekend_1, gen_last_weekend_2] if r]
                
                # Find first weekend of current month
                first_weekend_days = set()
                for day in range(1, scheduler.days_in_month + 1):
                    if scheduler.get_day_type(day) in ['thu', 'fri', 'sat']:
                        first_weekend_days.add(day)
                        if len(first_weekend_days) >= 3:
                            break
                
                for rad in last_weekend_gen:
                    for day in first_weekend_days:
                        scheduler.special_requests_off['GEN'].add((rad, day))
                
                # Section 2: GEN last day
                if gen_last_day:
                    scheduler.special_requests_off['GEN'].add((gen_last_day, 1))
                    scheduler.special_requests_off['GEN'].add((gen_last_day, 2))
                
                # Section 3: IRA last weekend
                if ira_last_weekend:
                    for day in range(1, min(8, scheduler.days_in_month + 1)):
                        scheduler.soft_constraints_off['IRA'].add((ira_last_weekend, day))
                
                # Section 4: Additional requests
                for req in st.session_state.additional_requests:
                    if req['day'] <= scheduler.days_in_month:
                        if req['hard']:
                            scheduler.special_requests_off[req['section']].add((req['rad'], req['day']))
                        else:
                            scheduler.soft_constraints_off[req['section']].add((req['rad'], req['day']))
                
                # Generate the schedule
                output_path = scheduler.generate_schedule()

                # Get quality metrics from MRI assignment
                quality_metrics = getattr(scheduler, 'mri_quality_metrics', None)
                
                # Display MRI Assignment Quality
                st.markdown("---")
                st.subheader("üìä MRI Assignment Quality Assessment")
                
                if quality_metrics:
                    # Create columns for metrics
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric(
                            "2-Rad Days", 
                            quality_metrics['two_rad_days'],
                            help="Days handled by Excel formulas (GEN or IRA can do MRI)"
                        )
                    
                    with col2:
                        st.metric(
                            "3-Rad Days", 
                            quality_metrics['three_rad_days'],
                            delta="Lower is better",
                            delta_color="inverse"
                        )
                    
                    with col3:
                        st.metric(
                            "Optimization Level",
                            quality_metrics['optimization_level'].replace('‚úì', '').replace('‚ö†', '').strip()
                        )
                    
                    # Show breakdown with weekend emphasis
                    if quality_metrics['three_rad_days'] > 0:
                        st.markdown("**3-Rad Day Breakdown:**")
                        col1, col2 = st.columns(2)
                        
                        weekend_count = quality_metrics['three_rad_weekends']
                        weekday_count = quality_metrics['three_rad_weekdays']
                        
                        with col1:
                            if weekend_count > 0:
                                waste_days = weekend_count * 3
                                st.error(f"üî¥ Weekends: {weekend_count} triplets ({waste_days} days wasted!)")
                                st.caption("‚ö†Ô∏è CRITICAL: MRI rad doing nothing but MRI")
                            else:
                                st.success(f"‚úÖ Weekends: 0 triplets (optimal!)")
                        
                        with col2:
                            if weekday_count > 0:
                                st.warning(f"üü° Weekdays: {weekday_count} days")
                            else:
                                st.success(f"‚úÖ Weekdays: 0 days (optimal!)")
                        
                        # Calculate total wasted capacity
                        wasted_weekend_days = quality_metrics['three_rad_weekends'] * 3
                        total_wasted = wasted_weekend_days + quality_metrics['three_rad_weekdays']
                        
                        if total_wasted > 0:
                            st.markdown("---")
                            st.markdown("**Capacity Impact:**")
                            
                            if wasted_weekend_days > 0:
                                st.error(
                                    f"‚ö†Ô∏è **{wasted_weekend_days} weekend days wasted** "
                                    f"({quality_metrics['three_rad_weekends']} triplets √ó 3 days)"
                                )
                                st.caption(
                                    "These are MRI rads doing ONLY MRI coverage without any GEN/IRA work. "
                                    "Goal: Zero 3-rad weekends!"
                                )
                            
                            if quality_metrics['three_rad_weekdays'] > 0:
                                st.info(
                                    f"‚ÑπÔ∏è {quality_metrics['three_rad_weekdays']} weekday assignments "
                                    f"could be optimized"
                                )
                        
                        # Show MRI-only distribution if available
                        if quality_metrics.get('mri_only_distribution'):
                            with st.expander("üìã View MRI-Only Assignment Distribution"):
                                st.markdown("**Rads with MRI-only assignments (not on GEN/IRA):**")
                                
                                for rad, counts in quality_metrics['mri_only_distribution'].items():
                                    weekend_count = counts['weekend_triplets']
                                    weekday_count = counts['weekdays']
                                    
                                    weekend_status = "‚úì" if weekend_count <= 1 else "‚ö† Over limit"
                                    weekday_status = "‚úì" if weekday_count <= 2 else "‚ö† Over limit"
                                    
                                    st.markdown(
                                        f"- **{rad}**: {weekend_count} weekend triplets {weekend_status}, "
                                        f"{weekday_count} weekdays {weekday_status}"
                                    )
                    else:
                        st.success("‚úÖ Perfect optimization! All MRI assignments handled by Excel formulas.")
                    
                    # Quality interpretation
                    three_rad_count = quality_metrics['three_rad_days']
                    if three_rad_count == 0:
                        st.success("üéâ **PERFECTLY OPTIMIZED**: Zero 3-rad days!")
                    elif three_rad_count <= 3:
                        st.success("‚úÖ **WELL OPTIMIZED**: Minimal 3-rad days")
                    elif three_rad_count <= 6:
                        st.warning("‚ö†Ô∏è **MODERATELY OPTIMIZED**: Could be improved")
                    else:
                        st.error("‚ö†Ô∏è **POORLY OPTIMIZED**: Needs significant improvement")
                else:
                    st.warning("‚ö†Ô∏è Quality metrics not available. Using older scheduler version?")
                
                # Calculate and display YTD Variance
                st.markdown("---")
                st.subheader("üìà YTD Variance Analysis")
                
                with st.spinner("Calculating YTD variance..."):
                    # Read the generated file to get YTD values
                    output_wb = openpyxl.load_workbook(output_path, data_only=True)
                    output_ws = output_wb['Sheet1']
                    
                    # Calculate variance
                    variance_results = scheduler.calculate_ytd_variance(output_ws)
                    
                    # Display summary metrics
                    col1, col2, col3, col4 = st.columns(4)
                    
                    summary = variance_results['summary']
                    
                    with col1:
                        weekend_rmse = summary['weekend']['rmse']
                        weekend_color = "normal" if weekend_rmse < 1.5 else "inverse"
                        st.metric(
                            "Weekend RMSE",
                            f"{weekend_rmse:.2f}",
                            delta="Lower is better",
                            delta_color=weekend_color
                        )
                    
                    with col2:
                        thu_rmse = summary['thu']['rmse']
                        thu_color = "normal" if thu_rmse < 1.5 else "inverse"
                        st.metric(
                            "Thursday RMSE",
                            f"{thu_rmse:.2f}",
                            delta="Lower is better",
                            delta_color=thu_color
                        )
                    
                    with col3:
                        weekday_rmse = summary['weekday']['rmse']
                        weekday_color = "normal" if weekday_rmse < 1.5 else "inverse"
                        st.metric(
                            "Weekday RMSE",
                            f"{weekday_rmse:.2f}",
                            delta="Lower is better",
                            delta_color=weekday_color
                        )
                    
                    with col4:
                        overall_score = variance_results['overall_score']
                        overall_color = "normal" if overall_score < 2.0 else "inverse"
                        st.metric(
                            "Overall Score",
                            f"{overall_score:.2f}",
                            delta="Weighted RMSE",
                            delta_color=overall_color
                        )
                    
                    # Quality interpretation
                    if overall_score < 1.5:
                        st.success("‚úÖ **EXCELLENT** balance between actual and target YTD!")
                    elif overall_score < 2.5:
                        st.success("‚úÖ **GOOD** balance")
                    elif overall_score < 3.5:
                        st.warning("‚ö†Ô∏è **FAIR** balance - could be improved")
                    else:
                        st.error("‚ö†Ô∏è **POOR** balance - needs improvement")
                    
                    # Detailed variance table
                    with st.expander("üìä View Detailed Variance Data"):
                        st.markdown("**Aggregate Variance by Day Type:**")
                        st.markdown("(Total Absolute Variance across all rads)")
                        
                        import pandas as pd
                        
                        variance_df = pd.DataFrame([
                            {
                                'Day Type': 'Weekend',
                                'Total Abs Variance': f"{summary['weekend']['total_abs_variance']:.2f}",
                                'Avg Abs Variance': f"{summary['weekend']['avg_abs_variance']:.3f}",
                                'RMSE': f"{summary['weekend']['rmse']:.3f}",
                                'Status': '‚úì' if summary['weekend']['rmse'] < 1.5 else '‚ö†'
                            },
                            {
                                'Day Type': 'Thursday',
                                'Total Abs Variance': f"{summary['thu']['total_abs_variance']:.2f}",
                                'Avg Abs Variance': f"{summary['thu']['avg_abs_variance']:.3f}",
                                'RMSE': f"{summary['thu']['rmse']:.3f}",
                                'Status': '‚úì' if summary['thu']['rmse'] < 1.5 else '‚ö†'
                            },
                            {
                                'Day Type': 'Weekday',
                                'Total Abs Variance': f"{summary['weekday']['total_abs_variance']:.2f}",
                                'Avg Abs Variance': f"{summary['weekday']['avg_abs_variance']:.3f}",
                                'RMSE': f"{summary['weekday']['rmse']:.3f}",
                                'Status': '‚úì' if summary['weekday']['rmse'] < 1.5 else '‚ö†'
                            }
                        ])
                        
                        st.dataframe(variance_df, use_container_width=True, hide_index=True)
                        
                        st.caption("**RMSE**: Root Mean Square Error - measures average deviation from target")
                        st.caption("**Overall Score**: Weighted RMSE (weekend√ó3 + thursday√ó2 + weekday√ó1) √∑ 6")
                
                st.markdown("---")
                
                
                # Display basic summary
                st.success("‚úÖ Schedule generated successfully!")
                
                month_name = calendar.month_name[scheduler.month]
                
                # Show assignment counts
                gen_count = len(scheduler.assignments['GEN'])
                ira_count = len(scheduler.assignments['IRA'])
                # Show assignment counts
                gen_count = len(scheduler.assignments['GEN'])
                ira_count = len(scheduler.assignments['IRA'])
                
                # Count MRI from quality metrics (includes both Python and formula assignments)
                if quality_metrics:
                    # Use 2-rad + 3-rad days for accurate count
                    mri_count = quality_metrics['two_rad_days'] + quality_metrics['three_rad_days']
                else:
                    # Fallback: assume all days covered
                    mri_count = scheduler.days_in_month
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("GEN Assignments", gen_count)
                with col2:
                    st.metric("IRA Assignments", ira_count)
                with col3:
                    st.metric("MRI Assignments", mri_count)
                
                # Check for completeness
                if gen_count == scheduler.days_in_month and ira_count == scheduler.days_in_month and mri_count == scheduler.days_in_month:
                    st.success("‚úÖ All days fully assigned!")
                else:
                    if gen_count != scheduler.days_in_month:
                        missing_gen = [d for d in range(1, scheduler.days_in_month + 1) if d not in scheduler.assignments['GEN']]
                        st.warning(f"‚ö†Ô∏è GEN missing days: {missing_gen}")
                    if ira_count != scheduler.days_in_month:
                        missing_ira = [d for d in range(1, scheduler.days_in_month + 1) if d not in scheduler.assignments['IRA']]
                        st.warning(f"‚ö†Ô∏è IRA missing days: {missing_ira}")
                    # MRI missing days check removed (MRI uses formulas for 2-rad days)
                    if mri_count != scheduler.days_in_month:
                        st.info(f"‚ÑπÔ∏è MRI: {scheduler.days_in_month - mri_count} days use formula-based assignment (verify by opening in Excel)")
                
                # Download button
                with open(output_path, 'rb') as f:
                    excel_data = f.read()
                
                filename = f"OnCall_Schedule_{month_name}_{scheduler.year}_GENERATED.xlsx"
                st.download_button(
                    label="üì• Download Generated Schedule",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                
        except Exception as e:
            st.error(f"‚ùå Error: {str(e)}")
            with st.expander("üîç View Error Details"):
                import traceback
                st.code(traceback.format_exc())

st.markdown("---")

# Section 3: Convert Rad Schedules for Import (Functional)
st.header("üîÑ Convert Rad Schedules for Import")
st.markdown("Upload completed Work Schedule and On-Call Schedule files to generate the import CSV file.")

# File uploaders
st.subheader("Upload Schedule Files (in order):")

col1, col2 = st.columns(2)

with col1:
    work_file = st.file_uploader(
        "1. Work Schedule (Excel)", 
        type=['xlsx'],
        key='work_schedule',
        help="Upload the completed Work Schedule Excel file"
    )

with col2:
    oncall_file = st.file_uploader(
        "2. On-Call Schedule (Excel)", 
        type=['xlsx'],
        key='oncall_schedule',
        help="Upload the completed On-Call Schedule Excel file"
    )

# Process button and conversion logic
if work_file and oncall_file:
    st.markdown("---")
    
    # Add month/year selection
    st.subheader("üìÖ Select Processing Month")
    
    col1, col2 = st.columns(2)
    with col1:
        selected_month = st.selectbox(
            "Month",
            options=list(range(1, 13)),
            format_func=lambda x: calendar.month_name[x],
            index=10  # Default to November (index 10 for month 11)
        )
    
    with col2:
        selected_year = st.number_input(
            "Year",
            min_value=2020,
            max_value=2030,
            value=2025,
            step=1
        )
    
    st.info(f"üìÖ Will process: **{calendar.month_name[selected_month]} {selected_year}**")
    
    if st.button("üîÑ Convert to Import Format", type="primary"):
        try:
            with st.spinner("Processing schedules..."):
                # Save uploaded files temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_work:
                    tmp_work.write(work_file.getvalue())
                    work_path = tmp_work.name
                
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_oncall:
                    tmp_oncall.write(oncall_file.getvalue())
                    oncall_path = tmp_oncall.name
                
                # Load workbooks
                wb_work = openpyxl.load_workbook(work_path)
                ws_work = wb_work['WORK SCHEDULE']
                
                wb_oncall = openpyxl.load_workbook(oncall_path, data_only=True)
                ws_oncall = wb_oncall['Sheet1']
                
                # Process schedules with EXPLICIT month/year
                output_data = rad_converter.process_schedules(
                    ws_work, 
                    ws_oncall, 
                    selected_year,
                    selected_month
                )
                
                # Show success message
                st.success(f"‚úÖ Generated {len(output_data)} schedule entries")
                
                # Display team breakdown similar to Cardiology
                with st.expander("üìã View Details"):
                    month_name = calendar.month_name[selected_month]
                    st.markdown(f"**Month:** {month_name} {selected_year}")
                    st.markdown("")
                    st.markdown(f"**Total entries:** {len(output_data)}")
                    st.markdown("")
                    st.markdown("**Entries per team:**")
                    
                    # Count entries by team
                    team_counts = {}
                    for entry in output_data:
                        team = entry['TEAM']
                        team_counts[team] = team_counts.get(team, 0) + 1
                    
                    # Display with proper team names
                    team_display_names = {
                        'General Radiology': 'General Radiology (GEN)',
                        'Interventional Radiology': 'Interventional Radiology (IRA)',
                        'MRI': 'MRI'
                    }
                    
                    for team, count in sorted(team_counts.items()):
                        display_name = team_display_names.get(team, team)
                        st.markdown(f"‚Ä¢ {display_name}: {count} entries")
                
                # Create CSV output
                output = io.StringIO()
                fieldnames = ['EMPLOYEE', 'TEAM', 'STARTDATE', 'STARTTIME', 
                             'ENDDATE', 'ENDTIME', 'ROLE', 'NOTES', 'ORDER', 'TEAMCOMMENT']
                writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter='^')
                writer.writeheader()
                writer.writerows(output_data)
                
                csv_data = output.getvalue()
                
                # Provide download button
                st.download_button(
                    label="üì• Download Import_OnCall_Radiology.csv",
                    data=csv_data,
                    file_name="Epic_OnCall_Import_Radiology.csv",
                    mime="text/csv",
                    type="primary"
                )
                
                st.success("‚úÖ Conversion complete! Click the button above to download your file.")
                
                # Clean up temp files
                os.unlink(work_path)
                os.unlink(oncall_path)
                
        except Exception as e:
            st.error(f"‚ùå Error during conversion: {str(e)}")
            st.error("Please check that your Excel files have the correct structure and try again.")
            import traceback
            with st.expander("üîç View Error Details"):
                st.code(traceback.format_exc())

else:
    st.info("üëÜ Please upload both Excel files to begin conversion")

# Footer
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: gray; font-size: 0.8em;'>"
    "Schedule Management System | Powered by Streamlit"
    "</div>",
    unsafe_allow_html=True
)
