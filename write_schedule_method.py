# FIXED write_schedule_to_excel METHOD
# This version PRESERVES MRI formulas for 2-rad days and only writes to 3-rad days

def write_schedule_to_excel(self):
    """Write all assignments to Excel - preserving MRI formulas for 2-rad days"""
    import openpyxl
    from pathlib import Path
    import calendar
    
    print("\n=== Writing schedule to Excel ===")
    
    wb_write = openpyxl.load_workbook(self.excel_path)
    ws_write = wb_write['Sheet1']
    
    # Verify assignment counts before writing
    gen_days = len(self.assignments['GEN'])
    ira_days = len(self.assignments['IRA'])
    mri_days = len(self.assignments['MRI'])
    
    print(f"Assignment counts: GEN={gen_days}, IRA={ira_days}, MRI={mri_days} (3-rad days only)")
    print(f"Expected: {self.days_in_month} days for GEN and IRA")
    print(f"Note: MRI count shows only 3-rad days (2-rad days handled by formulas)")
    
    if gen_days != self.days_in_month:
        print(f"  WARNING: GEN has {gen_days} assignments (expected {self.days_in_month})")
    if ira_days != self.days_in_month:
        print(f"  WARNING: IRA has {ira_days} assignments (expected {self.days_in_month})")
    
    # Clear existing X marks for GEN and IRA sections only
    print("\n=== Clearing existing assignments (GEN and IRA only) ===")
    
    for day in range(1, self.days_in_month + 1):
        col = day + 3
        
        # Clear GEN section (skip TA row)
        for rad, row in self.GEN_ROWS.items():
            if rad == 'TA':
                continue
            # Only clear if not locked
            if not (day in self.locked_assignments['GEN'] and 
                   self.locked_assignments['GEN'][day] == rad):
                ws_write.cell(row, col).value = None
        
        # Clear IRA section
        for rad, row in self.IRA_ROWS.items():
            # Only clear if not locked
            if not (day in self.locked_assignments['IRA'] and 
                   self.locked_assignments['IRA'][day] == rad):
                ws_write.cell(row, col).value = None
    
    print("✓ GEN and IRA sections cleared (MRI formulas preserved)")
    
    # Write GEN assignments
    print("\n=== Writing GEN assignments ===")
    for day, rad in self.assignments['GEN'].items():
        if day > self.days_in_month:
            print(f"  ERROR: Skipping GEN day {day} (beyond month end)")
            continue
        row = self.GEN_ROWS[rad]
        col = day + 3
        ws_write.cell(row, col, 'X')
    print(f"✓ Wrote {len(self.assignments['GEN'])} GEN assignments")
    
    # Write IRA assignments
    print("\n=== Writing IRA assignments ===")
    for day, rad in self.assignments['IRA'].items():
        if day > self.days_in_month:
            print(f"  ERROR: Skipping IRA day {day} (beyond month end)")
            continue
        row = self.IRA_ROWS[rad]
        col = day + 3
        ws_write.cell(row, col, 'X')
    print(f"✓ Wrote {len(self.assignments['IRA'])} IRA assignments")
    
    # Handle MRI assignments - CRITICAL: Only modify 3-rad days
    print("\n=== Writing MRI assignments (3-rad days only) ===")
    
    if len(self.assignments['MRI']) == 0:
        print("✓ No 3-rad days! All MRI assignments handled by Excel formulas")
    else:
        # Identify which days are 3-rad days by checking current MRI assignments
        three_rad_days = set(self.assignments['MRI'].keys())
        
        print(f"Identified {len(three_rad_days)} 3-rad days: {sorted(three_rad_days)}")
        
        # ONLY clear MRI cells for 3-rad days (to override formulas)
        # DO NOT touch MRI cells for 2-rad days (preserve formulas)
        for day in three_rad_days:
            col = day + 3
            # Clear ALL MRI rads for this specific day only
            for rad, row in self.MRI_ROWS.items():
                # Only clear if not locked
                if not (day in self.locked_assignments['MRI'] and 
                       self.locked_assignments['MRI'][day] == rad):
                    cell = ws_write.cell(row, col)
                    # Check if this cell has a formula
                    if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                        print(f"  Day {day}, Rad {rad}: Clearing formula '{cell.value}'")
                    cell.value = None
        
        # Now write the 3-rad day assignments
        for day, rad in self.assignments['MRI'].items():
            if day > self.days_in_month:
                print(f"  ERROR: Skipping MRI day {day} (beyond month end)")
                continue
            row = self.MRI_ROWS[rad]
            col = day + 3
            ws_write.cell(row, col, 'X')
            print(f"  Day {day} -> {rad} (3-rad day, overriding formula)")
        
        print(f"✓ Wrote {len(self.assignments['MRI'])} MRI assignments for 3-rad days")
        print(f"✓ Preserved MRI formulas for {self.days_in_month - len(three_rad_days)} 2-rad days")
    
    # Save output file
    base_name = Path(self.excel_path).stem
    month_name = calendar.month_name[self.month]
    output_path = Path(self.excel_path).parent / f"{base_name}_COMPLETED_{month_name}_{self.year}.xlsx"
    
    wb_write.save(output_path)
    print(f"\n✓ Schedule saved to: {output_path}")
    
    return str(output_path)
