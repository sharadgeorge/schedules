"""
OnCall Schedule Generator for Radiology Department - Enhanced Streamlit Version
Compatible with Streamlit web interface - no interactive prompts

Key Features:
- Dynamic YTD target calculations based on start dates
- Improved weekend triplet consolidation for partial weekends
- YTD caching system to prevent double-counting
- Intelligent load balancing with monthly limits
- Natural balancing: High YTD → fewer assignments, Low YTD → more assignments
- NO INTERACTIVE PROMPTS - preferences set via attributes
"""

import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import calendar

# Configuration
MAX_MONTHLY_GEN = 5  # Adjustable
MAX_MONTHLY_IRA = 12  # Adjustable
MAX_MONTHLY_MRI = 5  # Adjustable
MAX_MONTHLY_WEEKENDS_GEN = 3  # Hard limit for GEN weekends only
CONSECUTIVE_WEEKEND_PENALTY = 100
SOFT_CONSTRAINT_PENALTY = 300

# Employee start dates (month, day)
START_DATES = {
    'NN': (7, 1), 'MB': (1, 1), 'LK': (1, 1), 'PR': (7, 1),
    'AT': (1, 1), 'AK': (1, 1), 'MC': (1, 1), 'AO': (1, 1),
    'MM': (1, 1), 'IG': (5, 1), 'MF': (1, 1), 'AS': (1, 1)
}

# Section definitions
GEN_RADS = ['NN', 'MB', 'LK', 'PR', 'AT', 'AK', 'MC', 'AO', 'MM']
GEN_RADS_WITH_IRA = ['NN', 'MB', 'LK', 'PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'IG', 'MF', 'AS']
IRA_RADS = ['IG', 'MF', 'AS']
MRI_RADS = ['PR', 'AT', 'AK', 'MC', 'AO', 'MM', 'MF', 'AS']

# Excel row mappings
GEN_ROWS = {
    'TA': 5, 'NN': 7, 'MB': 8, 'LK': 9, 'PR': 10, 'AT': 11,
    'AK': 12, 'MC': 13, 'AO': 14, 'MM': 15, 'IG': 17, 'MF': 18, 'AS': 19
}
IRA_ROWS = {'IG': 24, 'MF': 25, 'AS': 26}
MRI_ROWS = {'PR': 30, 'AT': 31, 'AK': 32, 'MC': 33, 'AO': 34, 'MM': 35, 'MF': 36, 'AS': 37}

# Column indices (openpyxl uses 1-based indexing)
COL_AM = 39  # YTD weekday
COL_AN = 40  # YTD Thursday
COL_AO = 41  # YTD weekend
COL_AQ = 43  # YTD Target weekday
COL_AR = 44  # YTD Target Thursday
COL_AS = 45  # YTD Target weekend

# Holiday periods (month-start, day-start, month-end, day-end)
HOLIDAY_PERIODS = [
    (3, 27, 4, 5),   # Mar 27 - Apr 5
    (6, 5, 6, 14)    # Jun 5 - Jun 14
]


class OnCallScheduler:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.ws = self.wb['Sheet1']
        
        self.month, self.year = self.extract_month_year()
        self.days_in_month = calendar.monthrange(self.year, self.month)[1]
        
        # Cache YTD totals BEFORE any assignments
        self.ytd_cache = self.cache_ytd_totals()
        
        # Calculate YTD targets dynamically
        self.ytd_targets = self.calculate_ytd_targets()
        
        # Monthly counters
        self.monthly_counts = defaultdict(lambda: {'weekday': 0, 'thu': 0, 'weekend': 0})
        self.gen_monthly_total = defaultdict(int)
        self.ira_monthly_total = defaultdict(int)
        self.mri_monthly_total = defaultdict(int)
        
        # Assignments
        self.assignments = {'GEN': {}, 'IRA': {}, 'MRI': {}}
        
        # Constraints - initialized empty, set by Streamlit app
        self.locked_assignments = self.identify_locked_assignments()
        self.vacation_days = self.identify_vacation_days()
        self.special_requests_off = {}  # Hard constraints - SET BY STREAMLIT
        self.soft_constraints_off = {}   # Soft constraints - SET BY STREAMLIT


    def extract_month_year(self):
        """Extract month and year from Excel filename"""
        import re
        from pathlib import Path
        
        filename = Path(self.excel_path).stem
        
        months = {
            'january': 1, 'february': 2, 'march': 3, 'april': 4,
            'may': 5, 'june': 6, 'july': 7, 'august': 8,
            'september': 9, 'october': 10, 'november': 11, 'december': 12
        }
        
        for month_name, month_num in months.items():
            if month_name in filename.lower():
                year_match = re.search(r'20\d{2}', filename)
                if year_match:
                    return month_num, int(year_match.group())
        
        # Default to current date if not found
        today = datetime.now()
        return today.month, today.year


    def cache_ytd_totals(self):
        """Cache YTD totals from Excel BEFORE any assignments"""
        print("\n=== Caching YTD Totals ===")
        cache = {}
        
        # Cache GEN section
        for rad in GEN_RADS_WITH_IRA:
            if rad in GEN_ROWS:
                row = GEN_ROWS[rad]
                weekday = self.ws.cell(row, COL_AM).value or 0
                thu = self.ws.cell(row, COL_AN).value or 0
                weekend = self.ws.cell(row, COL_AO).value or 0
                
                cache[('GEN', rad)] = {
                    'weekday': int(weekday),
                    'thu': int(thu),
                    'weekend': int(weekend)
                }
                print(f"  GEN {rad}: WD={weekday}, Thu={thu}, WE={weekend}")
        
        # Cache IRA section
        for rad in IRA_RADS:
            if rad in IRA_ROWS:
                row = IRA_ROWS[rad]
                weekday = self.ws.cell(row, COL_AM).value or 0
                thu = self.ws.cell(row, COL_AN).value or 0
                weekend = self.ws.cell(row, COL_AO).value or 0
                
                cache[('IRA', rad)] = {
                    'weekday': int(weekday),
                    'thu': int(thu),
                    'weekend': int(weekend)
                }
                print(f"  IRA {rad}: WD={weekday}, Thu={thu}, WE={weekend}")
        
        # MRI section - no YTD tracking
        for rad in MRI_RADS:
            cache[('MRI', rad)] = {'weekday': 0, 'thu': 0, 'weekend': 0}
        
        print("✓ YTD totals cached successfully")
        return cache


    def calculate_availability_fraction(self, rad, day_type):
        """Calculate availability fraction for a rad based on start date"""
        start_month, start_day = START_DATES.get(rad, (1, 1))
        
        year_start = datetime(self.year, 1, 1)
        current_month_end = datetime(self.year, self.month, 
                                     calendar.monthrange(self.year, self.month)[1])
        rad_start = datetime(self.year, start_month, start_day)
        
        if rad_start > current_month_end:
            return 0.0
        
        # Count days available (excluding holidays)
        available_days = {'weekday': 0, 'thu': 0, 'weekend': 0}
        total_days = {'weekday': 0, 'thu': 0, 'weekend': 0}
        
        current_date = year_start
        while current_date <= current_month_end:
            if self.is_holiday(current_date):
                current_date += timedelta(days=1)
                continue
            
            day_of_week = current_date.weekday()
            
            if day_of_week == 3:  # Thursday
                dt = 'thu'
            elif day_of_week in [4, 5]:  # Friday or Saturday
                dt = 'weekend'
            else:
                dt = 'weekday'
            
            total_days[dt] += 1
            if current_date >= rad_start:
                available_days[dt] += 1
            
            current_date += timedelta(days=1)
        
        if total_days[day_type] == 0:
            return 0.0
        
        return available_days[day_type] / total_days[day_type]


    def is_holiday(self, date):
        """Check if a date falls within holiday periods"""
        for start_month, start_day, end_month, end_day in HOLIDAY_PERIODS:
            start_date = datetime(self.year, start_month, start_day)
            end_date = datetime(self.year, end_month, end_day)
            if start_date <= date <= end_date:
                return True
        return False


    def calculate_ytd_targets(self):
        """Calculate YTD targets dynamically based on start dates and availability"""
        print("\n=== Calculating YTD Targets ===")
        targets = {}
        
        # Count total days from Jan 1 to end of current month (excluding holidays)
        total_days = {'weekday': 0, 'thu': 0, 'weekend': 0}
        
        year_start = datetime(self.year, 1, 1)
        current_month_end = datetime(self.year, self.month, 
                                     calendar.monthrange(self.year, self.month)[1])
        
        current_date = year_start
        while current_date <= current_month_end:
            if not self.is_holiday(current_date):
                day_of_week = current_date.weekday()
                if day_of_week == 3:
                    total_days['thu'] += 1
                elif day_of_week in [4, 5]:
                    total_days['weekend'] += 1
                else:
                    total_days['weekday'] += 1
            current_date += timedelta(days=1)
        
        print(f"\nTotal days (Jan 1 - end of {calendar.month_name[self.month]}):")
        print(f"  Weekdays: {total_days['weekday']}")
        print(f"  Thursdays: {total_days['thu']}")
        print(f"  Weekends: {total_days['weekend']}")
        
        # Calculate targets for GEN section
        print("\n--- GEN Section Targets ---")
        
        total_availability = {'weekday': 0, 'thu': 0, 'weekend': 0}
        for rad in GEN_RADS:
            for day_type in ['weekday', 'thu', 'weekend']:
                frac = self.calculate_availability_fraction(rad, day_type)
                total_availability[day_type] += frac
        
        for rad in GEN_RADS:
            targets[('GEN', rad)] = {}
            for day_type in ['weekday', 'thu', 'weekend']:
                frac = self.calculate_availability_fraction(rad, day_type)
                if total_availability[day_type] > 0:
                    target = (total_days[day_type] / total_availability[day_type]) * frac
                else:
                    target = 0.0
                targets[('GEN', rad)][day_type] = target
            
            print(f"  {rad}: WD={targets[('GEN', rad)]['weekday']:.2f}, "
                  f"Thu={targets[('GEN', rad)]['thu']:.2f}, "
                  f"WE={targets[('GEN', rad)]['weekend']:.2f}")
        
        # Calculate targets for IRA section
        print("\n--- IRA Section Targets ---")
        
        total_availability_ira = {'weekday': 0, 'thu': 0, 'weekend': 0}
        for rad in IRA_RADS:
            for day_type in ['weekday', 'thu', 'weekend']:
                frac = self.calculate_availability_fraction(rad, day_type)
                total_availability_ira[day_type] += frac
        
        for rad in IRA_RADS:
            targets[('IRA', rad)] = {}
            for day_type in ['weekday', 'thu', 'weekend']:
                frac = self.calculate_availability_fraction(rad, day_type)
                if total_availability_ira[day_type] > 0:
                    target = (total_days[day_type] / total_availability_ira[day_type]) * frac
                else:
                    target = 0.0
                targets[('IRA', rad)][day_type] = target
            
            print(f"  {rad}: WD={targets[('IRA', rad)]['weekday']:.2f}, "
                  f"Thu={targets[('IRA', rad)]['thu']:.2f}, "
                  f"WE={targets[('IRA', rad)]['weekend']:.2f}")
        
        print("\n✓ YTD targets calculated")
        return targets


    def write_ytd_targets_to_excel(self):
        """Write calculated YTD targets to Excel columns AQ, AR, AS"""
        print("\n=== Writing YTD Targets to Excel ===")
        
        for rad in GEN_RADS:
            if ('GEN', rad) in self.ytd_targets:
                row = GEN_ROWS[rad]
                targets = self.ytd_targets[('GEN', rad)]
                
                self.ws.cell(row, COL_AQ, round(targets['weekday'], 2))
                self.ws.cell(row, COL_AR, round(targets['thu'], 2))
                self.ws.cell(row, COL_AS, round(targets['weekend'], 2))
        
        for rad in IRA_RADS:
            if ('IRA', rad) in self.ytd_targets:
                row = IRA_ROWS[rad]
                targets = self.ytd_targets[('IRA', rad)]
                
                self.ws.cell(row, COL_AQ, round(targets['weekday'], 2))
                self.ws.cell(row, COL_AR, round(targets['thu'], 2))
                self.ws.cell(row, COL_AS, round(targets['weekend'], 2))
        
        print("✓ YTD targets written to Excel")


    def identify_locked_assignments(self):
        """Identify pre-filled assignments (locked) in Excel"""
        locked = {'GEN': {}, 'IRA': {}, 'MRI': {}}
        
        for day in range(1, self.days_in_month + 1):
            col = day + 3
            
            for rad, row in GEN_ROWS.items():
                if rad == 'TA':
                    continue
                cell_value = self.ws.cell(row, col).value
                if cell_value == 'X':
                    locked['GEN'][day] = rad
            
            for rad, row in IRA_ROWS.items():
                cell_value = self.ws.cell(row, col).value
                if cell_value == 'X':
                    locked['IRA'][day] = rad
            
            for rad, row in MRI_ROWS.items():
                cell_value = self.ws.cell(row, col).value
                if cell_value == 'X':
                    locked['MRI'][day] = rad
        
        return locked


    def identify_vacation_days(self):
        """Identify vacation days marked in Excel"""
        vacation = defaultdict(set)
        
        for day in range(1, self.days_in_month + 1):
            col = day + 3
            
            for rad, row in {**GEN_ROWS, **IRA_ROWS, **MRI_ROWS}.items():
                if rad == 'TA':
                    continue
                cell_value = self.ws.cell(row, col).value
                if cell_value and str(cell_value).strip().upper() == 'V':
                    vacation[rad].add(day)
        
        return vacation


    def calculate_composite_score(self, section, rad, day_type):
        """Calculate composite score for assignment decision"""
        ytd = self.ytd_cache.get((section, rad), {}).get(day_type, 0)
        current_month = self.monthly_counts[rad][day_type]
        target = self.ytd_targets.get((section, rad), {}).get(day_type, 0)
        
        total = ytd + current_month
        
        if total < target:
            score = -(target - total)  # Negative = preferred
        else:
            score = (total - target) * 2  # Penalty
        
        return score, total, target


    def find_best_candidate(self, candidates, section, day_type, day):
        """Find best candidate using composite scoring"""
        if not candidates:
            return None, None
        
        best_rad = None
        best_score = float('inf')
        
        for rad in candidates:
            # Check monthly limits
            if section == 'GEN' and self.gen_monthly_total[rad] >= MAX_MONTHLY_GEN:
                continue
            if section == 'IRA' and self.ira_monthly_total[rad] >= MAX_MONTHLY_IRA:
                continue
            if section == 'MRI' and self.mri_monthly_total[rad] >= MAX_MONTHLY_MRI:
                continue
            
            # Check GEN weekend limit
            if section == 'GEN' and day_type == 'weekend':
                if self.monthly_counts[rad]['weekend'] >= MAX_MONTHLY_WEEKENDS_GEN:
                    continue
            
            # Calculate score
            score, total, target = self.calculate_composite_score(section, rad, day_type)
            
            # Apply soft constraint penalty
            if rad in self.soft_constraints_off and day in self.soft_constraints_off[rad]:
                score += SOFT_CONSTRAINT_PENALTY
            
            # Check consecutive weekend penalty
            if day_type == 'weekend':
                prev_weekend_days = self.get_previous_weekend_days(day)
                if any(self.assignments.get(section, {}).get(d) == rad for d in prev_weekend_days):
                    score += CONSECUTIVE_WEEKEND_PENALTY
            
            if score < best_score:
                best_score = score
                best_rad = rad
        
        return best_rad, best_score


    def get_previous_weekend_days(self, day):
        """Get days of previous weekend"""
        prev_days = []
        for d in range(max(1, day - 7), day):
            date = datetime(self.year, self.month, d)
            if date.weekday() in [4, 5]:
                prev_days.append(d)
        return prev_days


    def assign_gen(self):
        """Assign GEN call with intelligent load balancing"""
        print("\n=== Assigning GEN Call ===")
        
        for day in range(1, self.days_in_month + 1):
            if day in self.locked_assignments['GEN']:
                rad = self.locked_assignments['GEN'][day]
                self.assignments['GEN'][day] = rad
                print(f"Day {day}: {rad} (LOCKED)")
                continue
            
            date = datetime(self.year, self.month, day)
            day_of_week = date.weekday()
            
            if day_of_week == 3:
                day_type = 'thu'
            elif day_of_week in [4, 5]:
                day_type = 'weekend'
            else:
                day_type = 'weekday'
            
            # Get available candidates
            candidates = [rad for rad in GEN_RADS_WITH_IRA
                         if rad not in self.vacation_days or day not in self.vacation_days[rad]]
            
            # Remove hard constraints
            candidates = [rad for rad in candidates
                         if rad not in self.special_requests_off or 
                         day not in self.special_requests_off[rad]]
            
            # Thursday-Saturday pairing logic
            if day_of_week == 3:
                sat_day = day + 2
                if sat_day <= self.days_in_month:
                    paired_candidates = [rad for rad in candidates
                                        if sat_day not in self.locked_assignments['GEN']
                                        and (rad not in self.vacation_days or 
                                             sat_day not in self.vacation_days[rad])
                                        and (rad not in self.special_requests_off or
                                             sat_day not in self.special_requests_off[rad])]
                    
                    if paired_candidates:
                        rad, score = self.find_best_candidate(paired_candidates, 'GEN', 'thu', day)
                        if rad:
                            self.assignments['GEN'][day] = rad
                            self.monthly_counts[rad]['thu'] += 1
                            self.gen_monthly_total[rad] += 1
                            
                            self.assignments['GEN'][sat_day] = rad
                            self.monthly_counts[rad]['weekend'] += 1
                            self.gen_monthly_total[rad] += 1
                            
                            print(f"Day {day} (Thu) + Day {sat_day} (Sat): {rad} (PAIRED)")
                            continue
            
            # Normal assignment
            rad, score = self.find_best_candidate(candidates, 'GEN', day_type, day)
            
            if rad:
                self.assignments['GEN'][day] = rad
                self.monthly_counts[rad][day_type] += 1
                self.gen_monthly_total[rad] += 1
                print(f"Day {day} ({day_type}): {rad}")
            else:
                print(f"Day {day} ({day_type}): ERROR - No candidate found!")


    def assign_ira(self):
        """Assign IRA call with triplet logic"""
        print("\n=== Assigning IRA Call ===")
        
        for day in range(1, self.days_in_month + 1):
            if day in self.locked_assignments['IRA']:
                rad = self.locked_assignments['IRA'][day]
                self.assignments['IRA'][day] = rad
                print(f"Day {day}: {rad} (LOCKED)")
                continue
            
            date = datetime(self.year, self.month, day)
            day_of_week = date.weekday()
            
            if day_of_week == 3:
                day_type = 'thu'
            elif day_of_week in [4, 5]:
                day_type = 'weekend'
            else:
                day_type = 'weekday'
            
            candidates = [rad for rad in IRA_RADS
                         if rad not in self.vacation_days or day not in self.vacation_days[rad]]
            
            # Thursday triplet logic
            if day_of_week == 3:
                fri_day = day + 1
                sat_day = day + 2
                
                if fri_day <= self.days_in_month and sat_day <= self.days_in_month:
                    triplet_candidates = [rad for rad in candidates
                                         if (fri_day not in self.locked_assignments['IRA'] and
                                             sat_day not in self.locked_assignments['IRA'])
                                         and (rad not in self.vacation_days or
                                              (fri_day not in self.vacation_days[rad] and
                                               sat_day not in self.vacation_days[rad]))]
                    
                    if triplet_candidates:
                        rad, score = self.find_best_candidate(triplet_candidates, 'IRA', 'thu', day)
                        if rad:
                            self.assignments['IRA'][day] = rad
                            self.monthly_counts[rad]['thu'] += 1
                            self.ira_monthly_total[rad] += 1
                            
                            self.assignments['IRA'][fri_day] = rad
                            self.monthly_counts[rad]['weekend'] += 1
                            self.ira_monthly_total[rad] += 1
                            
                            self.assignments['IRA'][sat_day] = rad
                            self.monthly_counts[rad]['weekend'] += 1
                            self.ira_monthly_total[rad] += 1
                            
                            print(f"Day {day}-{sat_day} (Thu+Fri+Sat): {rad} (TRIPLET)")
                            continue
            
            # Normal assignment
            rad, score = self.find_best_candidate(candidates, 'IRA', day_type, day)
            
            if rad:
                self.assignments['IRA'][day] = rad
                self.monthly_counts[rad][day_type] += 1
                self.ira_monthly_total[rad] += 1
                print(f"Day {day} ({day_type}): {rad}")
            else:
                print(f"Day {day} ({day_type}): ERROR - No candidate found!")


    def assign_mri_3rad_days(self):
        """Assign MRI for 3-rad days with partial weekend consolidation"""
        print("\n=== Assigning MRI Call (3-rad days only) ===")
        
        thursdays = []
        for day in range(1, self.days_in_month + 1):
            date = datetime(self.year, self.month, day)
            if date.weekday() == 3:
                thursdays.append(day)
        
        print(f"Found {len(thursdays)} Thursdays: {thursdays}")
        
        for thu_day in thursdays:
            fri_day = thu_day + 1
            sat_day = thu_day + 2
            
            if sat_day > self.days_in_month:
                print(f"\nDay {thu_day} (Thu): Cannot form triplet (month ends early)")
                continue
            
            gen_thu = self.assignments['GEN'].get(thu_day)
            gen_fri = self.assignments['GEN'].get(fri_day)
            gen_sat = self.assignments['GEN'].get(sat_day)
            
            print(f"\nAnalyzing Thu {thu_day}, Fri {fri_day}, Sat {sat_day}:")
            print(f"  GEN assignments: Thu={gen_thu}, Fri={gen_fri}, Sat={gen_sat}")
            
            # Check for partial weekend consolidation
            mri_capable_gen_rads = set()
            if gen_thu and gen_thu in MRI_RADS:
                mri_capable_gen_rads.add(gen_thu)
            if gen_fri and gen_fri in MRI_RADS:
                mri_capable_gen_rads.add(gen_fri)
            if gen_sat and gen_sat in MRI_RADS:
                mri_capable_gen_rads.add(gen_sat)
            
            if len(mri_capable_gen_rads) == 1:
                rad = list(mri_capable_gen_rads)[0]
                days_on_gen = []
                if gen_thu == rad:
                    days_on_gen.append(thu_day)
                if gen_fri == rad:
                    days_on_gen.append(fri_day)
                if gen_sat == rad:
                    days_on_gen.append(sat_day)
                
                print(f"  → Partial weekend: {rad} on GEN for {len(days_on_gen)} days: {days_on_gen}")
                print(f"  → Consolidating: Assign {rad} to MRI for all 3 days")
                
                if self.mri_monthly_total[rad] + 3 <= MAX_MONTHLY_MRI:
                    if not (rad in self.vacation_days and 
                           any(d in self.vacation_days[rad] for d in [thu_day, fri_day, sat_day])):
                        self.assignments['MRI'][thu_day] = rad
                        self.assignments['MRI'][fri_day] = rad
                        self.assignments['MRI'][sat_day] = rad
                        self.mri_monthly_total[rad] += 3
                        print(f"  ✓ Assigned {rad} (partial weekend consolidation)")
                        continue
            
            # Check locked
            if (thu_day in self.locked_assignments['MRI'] or
                fri_day in self.locked_assignments['MRI'] or
                sat_day in self.locked_assignments['MRI']):
                locked_rad = (self.locked_assignments['MRI'].get(thu_day) or
                             self.locked_assignments['MRI'].get(fri_day) or
                             self.locked_assignments['MRI'].get(sat_day))
                print(f"  → LOCKED to {locked_rad}")
                self.assignments['MRI'][thu_day] = locked_rad
                self.assignments['MRI'][fri_day] = locked_rad
                self.assignments['MRI'][sat_day] = locked_rad
                continue
            
            # Find MRI-capable rads
            candidates = []
            for rad in MRI_RADS:
                if rad in self.vacation_days:
                    if any(d in self.vacation_days[rad] for d in [thu_day, fri_day, sat_day]):
                        continue
                
                if self.mri_monthly_total[rad] + 3 > MAX_MONTHLY_MRI:
                    continue
                
                candidates.append(rad)
            
            if not candidates:
                print(f"  ERROR: No MRI candidate available!")
                continue
            
            best_rad = min(candidates, key=lambda r: self.mri_monthly_total[r])
            
            self.assignments['MRI'][thu_day] = best_rad
            self.assignments['MRI'][fri_day] = best_rad
            self.assignments['MRI'][sat_day] = best_rad
            self.mri_monthly_total[best_rad] += 3
            
            print(f"  ✓ Assigned {best_rad} to MRI triplet")


    def generate_schedule(self):
        """Main scheduling workflow - NO INTERACTIVE PROMPTS"""
        print("\n" + "="*80)
        print(f"GENERATING SCHEDULE: {calendar.month_name[self.month]} {self.year}")
        print("="*80)
        
        # Assign sections
        self.assign_gen()
        self.assign_ira()
        self.assign_mri_3rad_days()
        
        # Write YTD targets
        self.write_ytd_targets_to_excel()
        
        # Write schedule
        output_path = self.write_schedule_to_excel()
        
        # Print summary
        self.print_summary()
        
        return output_path


    def print_summary(self):
        """Print assignment summary"""
        print("\n" + "="*80)
        print("ASSIGNMENT SUMMARY")
        print("="*80)
        
        print("\n--- Monthly Counts ---")
        for rad in sorted(set(GEN_RADS_WITH_IRA + IRA_RADS + MRI_RADS)):
            gen_count = self.gen_monthly_total[rad]
            ira_count = self.ira_monthly_total[rad]
            mri_count = self.mri_monthly_total[rad]
            
            if gen_count > 0 or ira_count > 0 or mri_count > 0:
                counts = self.monthly_counts[rad]
                print(f"{rad}: GEN={gen_count} (WD={counts['weekday']}, Thu={counts['thu']}, WE={counts['weekend']}), "
                      f"IRA={ira_count}, MRI={mri_count}")


    def write_schedule_to_excel(self):
        """Write schedule and targets to Excel"""
        from pathlib import Path
        
        print("\n=== Writing Schedule to Excel ===")
        
        # Clear and write GEN
        for day in range(1, self.days_in_month + 1):
            col = day + 3
            
            for rad, row in GEN_ROWS.items():
                if rad == 'TA':
                    continue
                if not (day in self.locked_assignments['GEN'] and 
                       self.locked_assignments['GEN'][day] == rad):
                    self.ws.cell(row, col).value = None
            
            if day in self.assignments['GEN']:
                rad = self.assignments['GEN'][day]
                row = GEN_ROWS[rad]
                self.ws.cell(row, col, 'X')
        
        print(f"✓ Wrote {len(self.assignments['GEN'])} GEN assignments")
        
        # Clear and write IRA
        for day in range(1, self.days_in_month + 1):
            col = day + 3
            
            for rad, row in IRA_ROWS.items():
                if not (day in self.locked_assignments['IRA'] and 
                       self.locked_assignments['IRA'][day] == rad):
                    self.ws.cell(row, col).value = None
            
            if day in self.assignments['IRA']:
                rad = self.assignments['IRA'][day]
                row = IRA_ROWS[rad]
                self.ws.cell(row, col, 'X')
        
        print(f"✓ Wrote {len(self.assignments['IRA'])} IRA assignments")
        
        # Write MRI (3-rad days only)
        if self.assignments['MRI']:
            three_rad_days = set(self.assignments['MRI'].keys())
            
            for day in three_rad_days:
                col = day + 3
                for rad, row in MRI_ROWS.items():
                    if not (day in self.locked_assignments['MRI'] and 
                           self.locked_assignments['MRI'][day] == rad):
                        self.ws.cell(row, col).value = None
            
            for day, rad in self.assignments['MRI'].items():
                row = MRI_ROWS[rad]
                col = day + 3
                self.ws.cell(row, col, 'X')
            
            print(f"✓ Wrote {len(self.assignments['MRI'])} MRI assignments (3-rad days)")
        else:
            print("✓ No 3-rad days (MRI handled by formulas)")
        
        # Save file
        base_name = Path(self.excel_path).stem
        month_name = calendar.month_name[self.month]
        output_filename = f"{base_name}_COMPLETED_{month_name}_{self.year}.xlsx"
        
        # For Streamlit, save to temp directory
        import tempfile
        output_path = Path(tempfile.gettempdir()) / output_filename
        
        self.wb.save(output_path)
        print(f"\n✓ Schedule saved to: {output_path}")
        
        return str(output_path)
