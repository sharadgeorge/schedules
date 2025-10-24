import streamlit as st
import sys
import os
from pathlib import Path
import tempfile
import calendar
from datetime import datetime, timedelta

# Configure page
st.set_page_config(
    page_title="Cardiology - Schedule Management",
    page_icon="❤",
    layout="centered",
    initial_sidebar_state="expanded"
)

# Sidebar with About section ONLY (no custom navigation)
with st.sidebar:
    st.markdown("### About")
    st.info("**Project by JA RAD**")

# Main page title - using HTML to force red color
st.markdown("# <span style='color: #e74c3c;'>❤</span> Cardiology Schedule Management", unsafe_allow_html=True)
st.markdown("---")

# Section: Convert Cardiology Schedules for Import
st.header("🔄 Convert Cardiology Schedules for Import")
st.markdown("Upload the required Cardiology Schedule Excel files to generate the import CSV file.")
st.info("ℹ️ Currently supports 2 input files. This will be expandable in the future for additional teams.")

# File uploaders
st.subheader("Upload Schedule Files (in order):")

col1, col2 = st.columns(2)

with col1:
    cardio_file1 = st.file_uploader(
        "1. Cardiovascular Schedule (Excel)", 
        type=['xlsx'],
        key='cardio_file1',
        help="Upload the Cardiovascular schedule Excel file"
    )

with col2:
    cardio_file2 = st.file_uploader(
        "2. Interventional Cardiologist Schedule (Excel)", 
        type=['xlsx'],
        key='cardio_file2',
        help="Upload the Interventional Cardiologist schedule Excel file"
    )

# Future expandability note
with st.expander("ℹ️ About Future Expansion"):
    st.markdown("""
    This converter currently processes:
    - **Team 8**: Cardiovascular (Echo Tech Adult & Pediatric)
    - **Team 94**: Interventional Cardiologist
    
    In the future, additional file upload slots can be added here for new teams.
    The converter script will need to be updated accordingly to handle the new team configurations.
    """)

# Process button and conversion logic
if cardio_file1 and cardio_file2:
    st.markdown("---")
    
    # Add month/year selection
    st.subheader("📅 Select Processing Month")
    
    col1, col2 = st.columns(2)
    with col1:
        selected_month = st.selectbox(
            "Month",
            options=list(range(1, 13)),
            format_func=lambda x: calendar.month_name[x],
            index=10,  # Default to November (index 10 for month 11)
            key='cardio_month'
        )
    
    with col2:
        selected_year = st.number_input(
            "Year",
            min_value=2020,
            max_value=2030,
            value=2025,
            step=1,
            key='cardio_year'
        )
    
    st.info(f"📅 Will process: **{calendar.month_name[selected_month]} {selected_year}**")
    
    if st.button("🔄 Convert to Import Format", type="primary"):
        try:
            with st.spinner("Processing Cardiology schedules..."):
                # Import the converter
                import oncall_converter_Cardiology_demo_v3 as cardio_converter
                import openpyxl
                import csv
                import io
                
                # Save uploaded files temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp1:
                    tmp1.write(cardio_file1.getvalue())
                    cardio_path1 = tmp1.name
                
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp2:
                    tmp2.write(cardio_file2.getvalue())
                    cardio_path2 = tmp2.name
                
                # Load workbooks
                wb_cardio1 = openpyxl.load_workbook(cardio_path1, data_only=True)
                wb_cardio2 = openpyxl.load_workbook(cardio_path2, data_only=True)
                
                # Debug: Show available worksheets
                st.info(f"📋 Cardiovascular file worksheets: {wb_cardio1.sheetnames}")
                st.info(f"📋 Interventional file worksheets: {wb_cardio2.sheetnames}")
                
                # Get worksheets - try specific names first
                if 'Nov-On call' in wb_cardio1.sheetnames:
                    ws_cardio1 = wb_cardio1['Nov-On call']
                elif 'Sheet1' in wb_cardio1.sheetnames:
                    ws_cardio1 = wb_cardio1['Sheet1']
                else:
                    ws_cardio1 = wb_cardio1[wb_cardio1.sheetnames[0]]
                
                st.success(f"✅ Using worksheet: '{ws_cardio1.title}' from Cardiovascular file")
                
                if 'Nov Attending' in wb_cardio2.sheetnames:
                    ws_cardio2 = wb_cardio2['Nov Attending']
                elif 'Sheet1' in wb_cardio2.sheetnames:
                    ws_cardio2 = wb_cardio2['Sheet1']
                else:
                    ws_cardio2 = wb_cardio2[wb_cardio2.sheetnames[0]]
                
                st.success(f"✅ Using worksheet: '{ws_cardio2.title}' from Interventional file")
                
                # Debug: Show sample data from worksheets
                with st.expander("🔍 Debug: View Sample Data"):
                    st.write("**Cardiovascular worksheet sample (rows 1-10, columns A-E):**")
                    cardio_sample = []
                    for row in range(1, min(11, ws_cardio1.max_row + 1)):
                        row_data = []
                        for col in range(1, 6):
                            cell = ws_cardio1.cell(row=row, column=col).value
                            row_data.append(str(cell) if cell else "")
                        cardio_sample.append(row_data)
                    st.table(cardio_sample)
                    
                    st.write("**Interventional worksheet sample (rows 1-10, columns A-E):**")
                    intv_sample = []
                    for row in range(1, min(11, ws_cardio2.max_row + 1)):
                        row_data = []
                        for col in range(1, 6):
                            cell = ws_cardio2.cell(row=row, column=col).value
                            row_data.append(str(cell) if cell else "")
                        intv_sample.append(row_data)
                    st.table(intv_sample)
                
                # Debug: Check what the converter expects
                st.info(f"🔍 Processing {calendar.month_name[selected_month]} {selected_year}")
                
                # Call converter functions
                cardio_data = cardio_converter.read_cardiovascular_data(
                    ws_cardio1,
                    selected_year,
                    selected_month
                )
                
                st.info(f"📊 Cardiovascular entries found: {len(cardio_data)}")
                
                intv_data = cardio_converter.read_interventional_data(
                    ws_cardio2,
                    selected_year,
                    selected_month
                )
                
                st.info(f"📊 Interventional entries found: {len(intv_data)}")
                
                # Combine data
                output_data = cardio_converter.create_output_data(
                    cardio_data,
                    intv_data
                )
                
                st.info(f"📊 Total combined entries: {len(output_data)}")
                
                # If no entries, provide troubleshooting help
                if len(output_data) == 0:
                    st.error("⚠️ No entries were generated!")
                    st.warning("""
                    **Possible reasons:**
                    1. Employee names in Excel don't match converter's employee mapping
                    2. No 'X' markers found in the day columns
                    3. Data is in different rows/columns than expected
                    4. Month/year in file doesn't match what you selected
                    
                    **Next steps:**
                    - Check the sample data above to see the file structure
                    - Verify employee names are in column A
                    - Verify 'X' markers are in the day columns
                    - Check that the data starts around row 5
                    """)
                    # Still allow download of empty file for inspection
                else:
                    st.success(f"✅ Generated {len(output_data)} schedule entries")
                
                # Show detailed statistics
                with st.expander("📊 View Detailed Statistics"):
                    st.write(f"**Month:** {calendar.month_name[selected_month]} {selected_year}")
                    st.write(f"**Total entries:** {len(output_data)}")
                    
                    # Team breakdown
                    team_counts = {}
                    for entry in output_data:
                        team = entry['TEAM']
                        team_counts[team] = team_counts.get(team, 0) + 1
                    
                    st.write("\n**Entries per team:**")
                    team_names = {
                        '8': 'Cardiovascular',
                        '94': 'Interventional Cardiologist'
                    }
                    for team_id in sorted(team_counts.keys()):
                        team_name = team_names.get(team_id, f"Team {team_id}")
                        st.write(f"- {team_name}: {team_counts[team_id]} entries")
                
                # Create CSV output
                output = io.StringIO()
                fieldnames = ['EMPLOYEE', 'TEAM', 'STARTDATE', 'STARTTIME', 
                             'ENDDATE', 'ENDTIME', 'ROLE', 'NOTES', 'ORDER', 'TEAMCOMMENT']
                writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter='^')
                writer.writeheader()
                writer.writerows(output_data)
                
                csv_data = output.getvalue()
                
                # Provide download button
                st.download_button(
                    label="📥 Download Import_OnCall_Cardiology.csv",
                    data=csv_data,
                    file_name="Import_OnCall_Cardiology.csv",
                    mime="text/csv",
                    type="primary"
                )
                
                st.success("✅ Conversion complete! Click the button above to download your file.")
                
                # Clean up temp files
                os.unlink(cardio_path1)
                os.unlink(cardio_path2)
                
        except Exception as e:
            st.error(f"❌ Error during conversion: {str(e)}")
            st.error("Please check that your Excel files have the correct structure and try again.")
            import traceback
            with st.expander("🔍 View Error Details"):
                st.code(traceback.format_exc())

else:
    st.info("👆 Please upload both Excel files to begin conversion")

# Footer
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: gray; font-size: 0.8em;'>"
    "Schedule Management System | Powered by Streamlit"
    "</div>",
    unsafe_allow_html=True
)
