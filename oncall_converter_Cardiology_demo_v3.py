import openpyxl
import csv
import calendar
import sys
from datetime import datetime, timedelta
from pathlib import Path

# Employee mapping dictionary - username is the key
EMPLOYEE_MAP = {
    'dosa0b': {'emp_initials': 'AG', 'emp_roles': ['2001'], 'emp_name': 'Anita Gunda'},
    'ghas4g': {'emp_initials': 'GS', 'emp_roles': ['84', '2001'], 'emp_name': 'Ghaitani S'},
    'rokas56': {'emp_initials': 'RK', 'emp_roles': ['84'], 'emp_name': 'R Kasturi'},
    'abherq': {'emp_initials': 'AE', 'emp_roles': ['84'], 'emp_name': 'Abe E M'},
    'villfh': {'emp_initials': 'VL', 'emp_roles': ['84'], 'emp_name': 'Village Lomba'},
    'qulfi6e': {'emp_initials': 'Q', 'emp_roles': ['3042457'], 'emp_name': 'Dr. Qureshi'},
    'sentri0': {'emp_initials': 'S', 'emp_roles': ['3042457'], 'emp_name': 'Dr. Bahri'},
}

# Create reverse lookups for flexible matching
INITIALS_TO_USERNAME = {v['emp_initials']: k for k, v in EMPLOYEE_MAP.items()}
NAME_TO_USERNAME = {v['emp_name']: k for k, v in EMPLOYEE_MAP.items()}

def find_username_by_identifier(identifier):
    """Find username by initials or name (flexible matching)"""
    if not identifier:
        return None
    
    identifier = str(identifier).strip()
    
    # Try exact match with initials (case-insensitive)
    identifier_upper = identifier.upper()
    if identifier_upper in INITIALS_TO_USERNAME:
        return INITIALS_TO_USERNAME[identifier_upper]
    
    # Try exact match with name
    if identifier in NAME_TO_USERNAME:
        return NAME_TO_USERNAME[identifier]
    
    # Try normalized name matching (without periods, case-insensitive)
    identifier_normalized = identifier.replace('.', '').strip()
    for name, username in NAME_TO_USERNAME.items():
        name_normalized = name.replace('.', '').strip()
        if identifier_normalized.lower() == name_normalized.lower():
            return username
    
    return None

# Mapping of cell markers to roles
MARKER_TO_ROLES = {
    'X': ['84', '2001'],    # Both Echo Tech Adult and Echo Tech Ped
    'XA': ['84'],           # Echo Tech Adult only
    'XP': ['2001'],         # Echo Tech Ped only
}

# Team configurations
TEAMS = {
    'Cardiovascular': {
        'team_id': '8',
        'file_index': 1,
        'sheet_cell': 'B4',
        'data_rows': (12, 16),
        'first_col': 'C',
        'last_col': 'AG'
    },
    'Interventional Cardiologist': {
        'team_id': '94',
        'file_index': 2,
        'sheet_cell': None,
        'data_rows': (31, 31),
        'first_col': 'D',
        'last_col': 'AH'
    },
}

def col_letter_to_index(col_letter):
    """Convert column letter to index (A=1, B=2, etc.)"""
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def is_weekday(date):
    """Check if date is a weekday (Sun-Thu, 0=Monday, 6=Sunday)"""
    day_of_week = date.weekday()
    # In Python: Monday=0, Tuesday=1, Wednesday=2, Thursday=3, Friday=4, Saturday=5, Sunday=6
    # Weekdays are: Sunday(6), Monday(0), Tuesday(1), Wednesday(2), Thursday(3)
    # Weekends are: Friday(4), Saturday(5)
    return day_of_week in [6, 0, 1, 2, 3]  # Sun, Mon, Tue, Wed, Thu

def extract_month_year_from_file(wb, sheet_name, cell_ref='B4'):
    """Extract month and year from Excel file"""
    try:
        ws = wb[sheet_name]
        cell_value = ws[cell_ref].value
        
        if isinstance(cell_value, datetime):
            return cell_value.month, cell_value.year
        
        # Try to parse string
        if isinstance(cell_value, str):
            cell_value = cell_value.strip().lower()
            # Try to find month name
            for month_idx in range(1, 13):
                month_name = calendar.month_name[month_idx].lower()
                month_abbr = calendar.month_abbr[month_idx].lower()
                if month_name in cell_value or month_abbr in cell_value:
                    # Look for year in nearby cells or use current year
                    year_cell = ws['D4'].value if 'D4' in ws else None
                    if year_cell and isinstance(year_cell, (int, float)):
                        return month_idx, int(year_cell)
                    # Try to find year in same cell
                    import re
                    year_match = re.search(r'20\d{2}', str(cell_value))
                    year = int(year_match.group()) if year_match else datetime.now().year
                    return month_idx, year
    except:
        pass
    
    return None, None

def get_sheet_for_month(wb, month_num):
    """Get the sheet corresponding to the specified month (month-agnostic)"""
    month_name = calendar.month_name[month_num]
    month_abbr = calendar.month_abbr[month_num]
    
    # Try to find sheet with month name
    for sheet_name in wb.sheetnames:
        sheet_lower = sheet_name.lower()
        if month_name.lower() in sheet_lower or month_abbr.lower() in sheet_lower:
            return wb[sheet_name]
    
    # If not found, return active sheet
    return wb.active

def read_cardiovascular_data(wb, month_num, year):
    """Read on-call data from Cardiovascular file"""
    config = TEAMS['Cardiovascular']
    ws = get_sheet_for_month(wb, month_num)
    
    # Get column range
    first_col_idx = col_letter_to_index(config['first_col'])
    last_col_idx = col_letter_to_index(config['last_col'])
    
    # Get row range
    row_start, row_end = config['data_rows']
    
    # Dictionary to store assignments: {day: [(username, [roles])]}
    assignments = {}
    days_in_month = calendar.monthrange(year, month_num)[1]
    
    # Read data for each day
    for day in range(1, days_in_month + 1):
        col_idx = first_col_idx + day - 1
        if col_idx > last_col_idx:
            break
        
        assignments[day] = []
        
        # Check each employee row
        for row in range(row_start, row_end + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            
            if cell_value:
                cell_str = str(cell_value).strip().upper()
                
                # Check if this is a valid marker
                if cell_str in MARKER_TO_ROLES:
                    # Get employee name/initials from column B (index 2)
                    emp_cell = ws.cell(row=row, column=2).value
                    if emp_cell:
                        username = find_username_by_identifier(emp_cell)
                        if username:
                            roles = MARKER_TO_ROLES[cell_str]
                            assignments[day].append((username, roles))
    
    return assignments

def read_interventional_data(wb, month_num, year):
    """Read on-call data from Interventional Cardiologist file"""
    config = TEAMS['Interventional Cardiologist']
    ws = get_sheet_for_month(wb, month_num)
    
    # Get column range
    first_col_idx = col_letter_to_index(config['first_col'])
    last_col_idx = col_letter_to_index(config['last_col'])
    
    # Get row
    row = config['data_rows'][0]
    
    # Dictionary to store assignments: {day: username}
    assignments = {}
    days_in_month = calendar.monthrange(year, month_num)[1]
    
    # Read data for each day
    for day in range(1, days_in_month + 1):
        col_idx = first_col_idx + day - 1
        if col_idx > last_col_idx:
            break
        
        cell_value = ws.cell(row=row, column=col_idx).value
        
        if cell_value:
            username = find_username_by_identifier(cell_value)
            if username:
                assignments[day] = username
    
    return assignments

def create_output_data(cardiovascular_data, interventional_data, year, month):
    """Create output data structure"""
    output_rows = []
    days_in_month = calendar.monthrange(year, month)[1]
    
    for day in range(1, days_in_month + 1):
        current_date = datetime(year, month, day)
        next_date = current_date + timedelta(days=1)
        
        # Format dates as M/D/YYYY without zero padding
        import platform
        if platform.system() == 'Windows':
            date_format = '%#m/%#d/%Y'
        else:
            date_format = '%-m/%-d/%Y'
        
        start_date_str = current_date.strftime(date_format)
        end_date_str = next_date.strftime(date_format)
        
        # Process Cardiovascular team (8)
        if day in cardiovascular_data:
            for username, roles in cardiovascular_data[day]:
                for role in roles:
                    row = {
                        'EMPLOYEE': username,
                        'TEAM': '8',
                        'STARTDATE': start_date_str,
                        'STARTTIME': '700',
                        'ENDDATE': end_date_str,
                        'ENDTIME': '700',
                        'ROLE': role,
                        'NOTES': '',
                        'ORDER': '',
                        'TEAMCOMMENT': ''
                    }
                    output_rows.append(row)
        
        # Process Interventional Cardiologist team (94)
        if day in interventional_data:
            username = interventional_data[day]
            
            # Determine start time based on day of week
            is_weekday_flag = is_weekday(current_date)
            start_time = '1600' if is_weekday_flag else '700'
            
            # Get role for this employee
            if username in EMPLOYEE_MAP:
                role = EMPLOYEE_MAP[username]['emp_roles'][0]
                
                row = {
                    'EMPLOYEE': username,
                    'TEAM': '94',
                    'STARTDATE': start_date_str,
                    'STARTTIME': start_time,
                    'ENDDATE': end_date_str,
                    'ENDTIME': '700',
                    'ROLE': role,
                    'NOTES': 'On Call',
                    'ORDER': '',
                    'TEAMCOMMENT': ''
                }
                output_rows.append(row)
    
    return output_rows

def write_output_files(output_data, output_dir, year, month):
    """Write output to both CSV (with caret delimiter) and Excel formats"""
    
    month_name = calendar.month_name[month]
    filename_prefix = "OnCall_Import_Cardiology"
    
    # Write CSV file with caret (^) delimiter
    csv_filename = output_dir / f"{filename_prefix}.csv"
    with open(csv_filename, 'w', newline='') as csvfile:
        fieldnames = ['EMPLOYEE', 'TEAM', 'STARTDATE', 'STARTTIME', 
                     'ENDDATE', 'ENDTIME', 'ROLE', 'NOTES', 'ORDER', 'TEAMCOMMENT']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter='^')
        writer.writeheader()
        writer.writerows(output_data)
    
    print(f"✓ Created CSV: {csv_filename}")
    
    # Write Excel file
    xlsx_filename = output_dir / f"{filename_prefix}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_name} OnCall"
    
    # Write headers
    headers = ['EMPLOYEE', 'TEAM', 'STARTDATE', 'STARTTIME', 
               'ENDDATE', 'ENDTIME', 'ROLE', 'NOTES', 'ORDER', 'TEAMCOMMENT']
    ws.append(headers)
    
    # Write data rows
    for row in output_data:
        ws.append([row[h] for h in headers])
    
    wb.save(xlsx_filename)
    print(f"✓ Created Excel: {xlsx_filename}")

def main():
    print("=" * 60)
    print("OnCall Converter - Cardiology v1")
    print("=" * 60)
    
    # Check if command line arguments were provided
    if len(sys.argv) >= 3:
        cardiovascular_file_path = sys.argv[1]
        interventional_file_path = sys.argv[2]
    else:
        # Prompt user for file paths
        print("\nPlease provide the input files:")
        cardiovascular_file_path = input("  Cardiovascular file (.xlsx): ").strip()
        interventional_file_path = input("  INTV_CARD file (.xlsx): ").strip()
        
        # Remove quotes if user copied path with quotes
        cardiovascular_file_path = cardiovascular_file_path.strip('"').strip("'")
        interventional_file_path = interventional_file_path.strip('"').strip("'")
    
    cardio_file = Path(cardiovascular_file_path)
    intv_file = Path(interventional_file_path)
    
    # Validate files exist
    if not cardio_file.exists():
        print(f"\n✗ Error: Cardiovascular file not found: {cardiovascular_file_path}")
        return
    
    if not intv_file.exists():
        print(f"\n✗ Error: INTV_CARD file not found: {interventional_file_path}")
        return
    
    # Output directory is the same as input files directory
    output_dir = cardio_file.parent
    
    print(f"\nLoading files...")
    print(f"  Cardiovascular: {cardio_file.name}")
    print(f"  Interventional: {intv_file.name}")
    
    # Load workbooks
    wb_cardio = openpyxl.load_workbook(cardio_file, data_only=True)
    wb_intv = openpyxl.load_workbook(intv_file, data_only=True)
    
    # Extract month and year from Cardiovascular file
    # Try to find the sheet with "On Call" or "On-Call" in the name
    cardio_sheet = None
    for sheet_name in wb_cardio.sheetnames:
        if 'on' in sheet_name.lower() and 'call' in sheet_name.lower():
            cardio_sheet = wb_cardio[sheet_name]
            break
    
    if not cardio_sheet:
        cardio_sheet = wb_cardio.active
    
    # Extract month from cell B4
    month_num, year = extract_month_year_from_file(wb_cardio, cardio_sheet.title, 'B4')
    
    if month_num is None:
        # Try to extract from filename
        filename_lower = cardio_file.stem.lower()
        for m in range(1, 13):
            if calendar.month_name[m].lower() in filename_lower:
                month_num = m
                break
        
        if month_num is None:
            print(f"\n⚠ Warning: Could not detect month, using current month")
            month_num = datetime.now().month
        
        if year is None:
            year = datetime.now().year
    
    month_name = calendar.month_name[month_num]
    print(f"\nProcessing: {month_name} {year}")
    print("-" * 60)
    
    # Read data from both files
    print("Reading Cardiovascular assignments...")
    cardiovascular_data = read_cardiovascular_data(wb_cardio, month_num, year)
    cardio_days_with_assignments = len([d for d in cardiovascular_data if cardiovascular_data[d]])
    print(f"  Found assignments on {cardio_days_with_assignments} days")
    
    print("Reading Interventional Cardiologist assignments...")
    interventional_data = read_interventional_data(wb_intv, month_num, year)
    intv_days_with_assignments = len(interventional_data)
    print(f"  Found assignments on {intv_days_with_assignments} days")
    
    # Create output data
    print("Generating output data...")
    output_data = create_output_data(cardiovascular_data, interventional_data, year, month_num)
    
    print(f"✓ Generated {len(output_data)} schedule entries")
    print(f"  Expected: {calendar.monthrange(year, month_num)[1] * 3} entries (3 per day)")
    
    # Show sample of first few entries
    if output_data:
        print("\nSample of first 6 entries:")
        for i, row in enumerate(output_data[:6]):
            print(f"  {i+1}. {row['EMPLOYEE']} | Team {row['TEAM']} | {row['STARTDATE']} {row['STARTTIME']}-{row['ENDTIME']} | Role {row['ROLE']}")
    
    # Write output files
    print("\nWriting output files...")
    write_output_files(output_data, output_dir, year, month_num)
    
    print(f"\n{'=' * 60}")
    print(f"✓ All files created successfully!")
    print(f"  Output directory: {output_dir}")
    print(f"{'=' * 60}")

if __name__ == "__main__":
    main()
